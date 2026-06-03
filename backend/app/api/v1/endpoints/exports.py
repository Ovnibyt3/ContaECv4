"""
ContaEC - Endpoints de Exportación a Excel/CSV/PDF/ZIP
Exportación de productos, clientes, comprobantes y kardex
a múltiples formatos: Excel (.xlsx), CSV, PDF y ZIP
"""
import csv
import io
import logging
import zipfile
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.client import Client
from app.models.company import Company
from app.models.comprobante import Comprobante
from app.models.kardex import Kardex
from app.models.product import Product
from app.models.user import User

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/exports", tags=["Exportaciones"])


# ==========================================
# Funciones auxiliares
# ==========================================

async def _get_company_for_user(
    db: AsyncSession,
    company_id: str,
    user_id: str,
) -> Company:
    """Obtiene una empresa verificando que pertenezca al usuario actual."""
    result = await db.execute(
        select(Company).where(
            Company.id == company_id,
            Company.user_id == user_id,
            Company.is_active == True,
        )
    )
    company = result.scalars().first()
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada o no pertenece al usuario actual.",
        )
    return company


def _get_excel_response(workbook_bytes: bytes, filename: str) -> StreamingResponse:
    """Crea una StreamingResponse para un archivo Excel."""
    return StreamingResponse(
        io.BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _get_csv_response(csv_content: str, filename: str) -> StreamingResponse:
    """Crea una StreamingResponse para un archivo CSV."""
    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8-sig")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _get_zip_response(zip_bytes: bytes, filename: str) -> StreamingResponse:
    """Crea una StreamingResponse para un archivo ZIP."""
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==========================================
# Exportación de Productos
# ==========================================

@router.get("/products/excel")
async def export_products_excel(
    company_id: str = Query(..., description="ID de la empresa"),
    is_active: bool | None = True,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exportar productos a archivo Excel (.xlsx)."""
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Product).where(Product.company_id == company_id)
    if is_active is not None:
        query = query.where(Product.is_active == is_active)
    query = query.order_by(Product.descripcion)

    result = await db.execute(query)
    products = result.scalars().all()

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Encabezados
        headers = [
            "Código Principal", "Código Auxiliar", "Descripción", "Tipo",
            "Precio Unitario", "IVA Código", "IVA %", "ICE Código", "ICE %",
            "Unidad Medida", "Descuento %", "Código Barras", "Stock",
            "Stock Mínimo", "Ubicación", "Activo",
        ]

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos
        for row_idx, product in enumerate(products, 2):
            values = [
                product.codigo_principal,
                product.codigo_auxiliar,
                product.descripcion,
                "Bien" if product.tipo == "B" else "Servicio",
                float(product.precio_unitario) if product.precio_unitario else 0,
                product.iva_codigo,
                float(product.iva_porcentaje) if product.iva_porcentaje else 0,
                product.ice_codigo,
                float(product.ice_porcentaje) if product.ice_porcentaje else None,
                product.unidad_medida,
                float(product.descuento) if product.descuento else 0,
                getattr(product, "codigo_barras", None),
                float(getattr(product, "stock", 0) or 0),
                float(getattr(product, "stock_minimo", 0) or 0),
                getattr(product, "ubicacion", None),
                "Sí" if product.is_active else "No",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Ajustar anchos de columna
        column_widths = [15, 15, 40, 10, 15, 12, 8, 12, 8, 15, 12, 18, 10, 12, 20, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return _get_excel_response(output.getvalue(), f"productos_{timestamp}.xlsx")

    except Exception as e:
        logger.error(f"Error al exportar Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo Excel: {str(e)}",
        )


@router.get("/products/csv")
async def export_products_csv(
    company_id: str = Query(..., description="ID de la empresa"),
    is_active: bool | None = True,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exportar productos a archivo CSV."""
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Product).where(Product.company_id == company_id)
    if is_active is not None:
        query = query.where(Product.is_active == is_active)
    query = query.order_by(Product.descripcion)

    result = await db.execute(query)
    products = result.scalars().all()

    try:
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        # Encabezados
        writer.writerow([
            "codigo_principal", "codigo_auxiliar", "descripcion", "tipo",
            "precio_unitario", "iva_codigo", "iva_porcentaje", "ice_codigo",
            "ice_porcentaje", "unidad_medida", "descuento", "codigo_barras",
            "stock", "stock_minimo", "ubicacion", "is_active",
        ])

        # Datos
        for product in products:
            writer.writerow([
                product.codigo_principal,
                product.codigo_auxiliar or "",
                product.descripcion,
                product.tipo,
                product.precio_unitario,
                product.iva_codigo,
                product.iva_porcentaje,
                product.ice_codigo or "",
                product.ice_porcentaje or "",
                product.unidad_medida,
                product.descuento,
                getattr(product, "codigo_barras", "") or "",
                getattr(product, "stock", 0) or 0,
                getattr(product, "stock_minimo", 0) or 0,
                getattr(product, "ubicacion", "") or "",
                product.is_active,
            ])

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return _get_csv_response(output.getvalue(), f"productos_{timestamp}.csv")

    except Exception as e:
        logger.error(f"Error al exportar CSV: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo CSV: {str(e)}",
        )


# ==========================================
# Exportación de Clientes
# ==========================================

@router.get("/clients/excel")
async def export_clients_excel(
    company_id: str = Query(..., description="ID de la empresa"),
    is_active: bool | None = True,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exportar clientes a archivo Excel (.xlsx)."""
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Client).where(Client.company_id == company_id)
    if is_active is not None:
        query = query.where(Client.is_active == is_active)
    query = query.order_by(Client.razon_social)

    result = await db.execute(query)
    clients = result.scalars().all()

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        headers = [
            "Tipo Identificación", "Identificación", "Razón Social",
            "Dirección", "Email", "Teléfono", "Consumidor Final", "Activo",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        tipo_nombres = {"04": "RUC", "05": "Cédula", "06": "Pasaporte",
                        "07": "Consumidor Final", "08": "Exterior"}

        for row_idx, client in enumerate(clients, 2):
            values = [
                tipo_nombres.get(client.tipo_identificacion, client.tipo_identificacion),
                client.identificacion,
                client.razon_social,
                client.direccion or "",
                client.email or "",
                client.telefono or "",
                "Sí" if client.is_default_consumer else "No",
                "Sí" if client.is_active else "No",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        column_widths = [18, 18, 35, 40, 30, 15, 16, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return _get_excel_response(output.getvalue(), f"clientes_{timestamp}.xlsx")

    except Exception as e:
        logger.error(f"Error al exportar Excel de clientes: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo Excel: {str(e)}",
        )


# ==========================================
# Exportación de Comprobantes
# ==========================================

@router.get("/comprobantes/excel")
async def export_comprobantes_excel(
    company_id: str = Query(..., description="ID de la empresa"),
    tipo_comprobante: str | None = None,
    estado: str | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exportar comprobantes a archivo Excel (.xlsx)."""
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Comprobante).where(Comprobante.company_id == company_id)
    if tipo_comprobante:
        query = query.where(Comprobante.tipo_comprobante == tipo_comprobante)
    if estado:
        query = query.where(Comprobante.estado == estado)
    query = query.order_by(Comprobante.fecha_emision.desc())

    result = await db.execute(query)
    comprobantes = result.scalars().all()

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comprobantes"

        headers = [
            "Tipo", "Secuencial", "Fecha Emisión", "Estado", "Clave Acceso",
            "Cliente Identificación", "Cliente Razón Social",
            "Subtotal Sin Impuestos", "Total IVA", "Total ICE",
            "Total Con Impuestos", "Forma Pago", "Autorización SRI",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        tipo_nombres = {
            "01": "Factura", "03": "Liquidación", "04": "Nota Crédito",
            "05": "Nota Débito", "06": "Guía Remisión", "07": "Retención",
        }

        for row_idx, comp in enumerate(comprobantes, 2):
            values = [
                tipo_nombres.get(comp.tipo_comprobante, comp.tipo_comprobante),
                comp.secuencial,
                comp.fecha_emision.strftime("%Y-%m-%d %H:%M") if comp.fecha_emision else "",
                comp.estado,
                comp.clave_acceso or "",
                comp.cliente_identificacion or "",
                comp.cliente_razon_social or "",
                float(comp.subtotal_sin_impuestos) if comp.subtotal_sin_impuestos else 0,
                float(comp.total_iva) if comp.total_iva else 0,
                float(comp.total_ice) if comp.total_ice else 0,
                float(comp.total_con_impuestos) if comp.total_con_impuestos else 0,
                comp.forma_pago,
                comp.numero_autorizacion or "",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        column_widths = [14, 14, 18, 14, 52, 18, 30, 18, 12, 12, 18, 12, 52]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return _get_excel_response(output.getvalue(), f"comprobantes_{timestamp}.xlsx")

    except Exception as e:
        logger.error(f"Error al exportar Excel de comprobantes: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo Excel: {str(e)}",
        )


@router.get("/comprobantes/pdf")
async def export_comprobantes_pdf(
    company_id: str = Query(..., description="ID de la empresa"),
    tipo_comprobante: str | None = None,
    estado: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exportar comprobantes a PDF (lote).

    Genera un PDF con un resumen/tabla de los comprobantes filtrados.
    Para RIDEs individuales usar /comprobantes/{id}/ride.
    """
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Comprobante).where(Comprobante.company_id == company_id)
    if tipo_comprobante:
        query = query.where(Comprobante.tipo_comprobante == tipo_comprobante)
    if estado:
        query = query.where(Comprobante.estado == estado)
    query = query.order_by(Comprobante.fecha_emision.desc())

    result = await db.execute(query)
    comprobantes = result.scalars().all()

    # Obtener datos de la empresa
    company = await _get_company_for_user(db, company_id, current_user.id)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontSize=16,
            spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=styles["Normal"],
            fontSize=10,
            spaceAfter=12,
        )

        elements = []

        # Título
        elements.append(Paragraph("Reporte de Comprobantes Electrónicos", title_style))
        elements.append(Paragraph(
            f"Empresa: {company.razon_social} | RUC: {company.ruc} | "
            f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            subtitle_style,
        ))
        elements.append(Spacer(1, 12))

        # Tabla
        tipo_nombres = {
            "01": "Factura", "03": "Liquidación", "04": "N. Crédito",
            "05": "N. Débito", "06": "Guía", "07": "Retención",
        }

        table_data = [[
            "Tipo", "Secuencial", "Fecha", "Estado", "Cliente",
            "Subtotal", "IVA", "Total", "Autorización",
        ]]

        for comp in comprobantes:
            table_data.append([
                tipo_nombres.get(comp.tipo_comprobante, comp.tipo_comprobante),
                comp.secuencial,
                comp.fecha_emision.strftime("%Y-%m-%d") if comp.fecha_emision else "",
                comp.estado,
                comp.cliente_razon_social or "N/A",
                f"${comp.subtotal_sin_impuestos:.2f}" if comp.subtotal_sin_impuestos else "$0.00",
                f"${comp.total_iva:.2f}" if comp.total_iva else "$0.00",
                f"${comp.total_con_impuestos:.2f}" if comp.total_con_impuestos else "$0.00",
                comp.numero_autorizacion or "",
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitespace),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        elements.append(table)
        doc.build(elements)

        output.seek(0)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=comprobantes_{timestamp}.pdf"},
        )

    except Exception as e:
        logger.error(f"Error al exportar PDF: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo PDF: {str(e)}",
        )


# ==========================================
# Exportación de Kardex
# ==========================================

@router.get("/kardex/excel")
async def export_kardex_excel(
    company_id: str = Query(..., description="ID de la empresa"),
    product_id: str | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exportar movimientos de kardex a archivo Excel (.xlsx)."""
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Kardex).where(
        Kardex.company_id == company_id,
        Kardex.is_active == True,
    )
    if product_id:
        query = query.where(Kardex.product_id == product_id)
    query = query.order_by(Kardex.fecha_movimiento.asc())

    result = await db.execute(query)
    movements = result.scalars().all()

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kardex"

        headers = [
            "Fecha", "Tipo Movimiento", "Producto ID", "Cantidad",
            "Costo Unitario", "Costo Total", "Saldo Cantidad",
            "Saldo Valor", "Referencia Tipo", "Referencia ID",
            "Referencia Secuencial", "Detalle",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        tipo_nombres = {"entrada": "Entrada", "salida": "Salida", "ajuste": "Ajuste"}

        for row_idx, mov in enumerate(movements, 2):
            values = [
                mov.fecha_movimiento.strftime("%Y-%m-%d %H:%M") if mov.fecha_movimiento else "",
                tipo_nombres.get(mov.tipo_movimiento, mov.tipo_movimiento),
                mov.product_id,
                float(mov.cantidad) if mov.cantidad else 0,
                float(mov.costo_unitario) if mov.costo_unitario else 0,
                float(mov.costo_total) if mov.costo_total else 0,
                float(mov.saldo_cantidad) if mov.saldo_cantidad else 0,
                float(mov.saldo_valor) if mov.saldo_valor else 0,
                mov.referencia_tipo or "",
                mov.referencia_id or "",
                mov.referencia_secuencial or "",
                mov.detalle or "",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        column_widths = [18, 16, 36, 12, 14, 14, 14, 14, 16, 36, 16, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return _get_excel_response(output.getvalue(), f"kardex_{timestamp}.xlsx")

    except Exception as e:
        logger.error(f"Error al exportar Excel de kardex: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo Excel: {str(e)}",
        )


# ==========================================
# Exportación de XML como ZIP
# ==========================================

@router.get("/comprobantes/xml-zip")
async def export_comprobantes_xml_zip(
    company_id: str = Query(..., description="ID de la empresa"),
    estado: str | None = None,
    tipo_comprobante: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exportar archivos XML de comprobantes como ZIP.

    Incluye todos los XML firmados/autorizados de los comprobantes filtrados.
    """
    await _get_company_for_user(db, company_id, current_user.id)

    query = select(Comprobante).where(
        Comprobante.company_id == company_id,
        Comprobante.xml_content.isnot(None),  # type: ignore
    )
    if estado:
        query = query.where(Comprobante.estado == estado)
    if tipo_comprobante:
        query = query.where(Comprobante.tipo_comprobante == tipo_comprobante)
    query = query.order_by(Comprobante.fecha_emision.desc())

    result = await db.execute(query)
    comprobantes = result.scalars().all()

    if not comprobantes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontraron comprobantes con XML para exportar.",
        )

    try:
        tipo_nombres = {
            "01": "factura", "03": "liquidacion", "04": "nota_credito",
            "05": "nota_debito", "06": "guia_remision", "07": "retencion",
        }

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for comp in comprobantes:
                if comp.xml_content:
                    tipo = tipo_nombres.get(comp.tipo_comprobante, "comprobante")
                    secuencial = comp.secuencial or comp.id[:8]
                    filename = f"{tipo}_{secuencial}.xml"
                    zipf.writestr(filename, comp.xml_content)

        zip_buffer.seek(0)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return _get_zip_response(zip_buffer.getvalue(), f"xml_comprobantes_{timestamp}.zip")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al exportar ZIP de XMLs: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo ZIP: {str(e)}",
        )
