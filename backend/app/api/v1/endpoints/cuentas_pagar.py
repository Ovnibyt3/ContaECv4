"""
ContaEC - Endpoints de Cuentas por Pagar

Gestión de cuentas por pagar a proveedores: listado, pagos, renegociación,
resúmenes y exportación a Excel/CSV.
"""
import csv
import io
import logging
from datetime import datetime, timedelta, timezone
from decimal import Decimal

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_action
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.company import Company
from app.models.purchase import CuentaPorPagar
from app.models.supplier import Supplier
from app.models.user import User
from app.schemas.purchase import (
    CuentaPorPagarPayment,
    CuentaPorPagarRenegotiation,
    CuentaPorPagarResponse,
    CuentaPorPagarSummary,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/cuentas-pagar", tags=["Cuentas por Pagar"])


# ==========================================
# Funciones auxiliares
# ==========================================

async def _get_company_for_user(
    db: AsyncSession,
    company_id: str,
    user_id: str,
) -> Company:
    """Obtiene una empresa verificando que pertenezca al usuario actual."""
    result = await db.execute(
        select(Company).where(
            Company.id == company_id,
            Company.user_id == user_id,
            Company.is_active == True,
        )
    )
    company = result.scalars().first()
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada o no pertenece al usuario actual.",
        )
    return company


async def _get_cuenta_or_404(
    db: AsyncSession,
    cuenta_id: str,
    current_user: User,
) -> CuentaPorPagar:
    """Obtiene una cuenta por pagar verificando que pertenezca a una empresa del usuario."""
    result = await db.execute(
        select(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(
            CuentaPorPagar.id == cuenta_id,
            Company.user_id == current_user.id,
        )
    )
    cuenta = result.scalars().first()
    if not cuenta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cuenta por pagar no encontrada.",
        )
    return cuenta


def _check_overdue(cuenta: CuentaPorPagar) -> str:
    """Verifica y actualiza el estado a 'vencida' si corresponde. Retorna el estado actualizado."""
    if (
        cuenta.estado in ("pendiente", "parcial")
        and cuenta.fecha_vencimiento
        and cuenta.fecha_vencimiento < datetime.now(timezone.utc)
    ):
        cuenta.estado = "vencida"
    return cuenta.estado


# ==========================================
# Listado y detalle
# ==========================================

@router.get("/", response_model=list[CuentaPorPagarResponse])
async def list_cuentas_por_pagar(
    company_id: str | None = Query(None, description="ID de la empresa"),
    supplier_id: str | None = Query(None, description="ID del proveedor"),
    estado: str | None = Query(None, description="Filtrar por estado"),
    vencimiento_proximo: int | None = Query(
        None,
        ge=1,
        le=365,
        description="Número de días para filtrar cuentas próximas a vencer",
    ),
    is_active: bool | None = Query(True, description="Solo activas"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Listar cuentas por pagar con filtros opcionales.

    - **company_id**: Filtrar por empresa
    - **supplier_id**: Filtrar por proveedor
    - **estado**: Filtrar por estado (pendiente, parcial, pagada, vencida, anulada)
    - **vencimiento_proximo**: Días dentro de los cuales vence la cuenta
    """
    query = (
        select(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(CuentaPorPagar.company_id == company_id)
    if supplier_id:
        query = query.where(CuentaPorPagar.supplier_id == supplier_id)
    if estado:
        query = query.where(CuentaPorPagar.estado == estado)
    if is_active is not None:
        query = query.where(CuentaPorPagar.is_active == is_active)
    if vencimiento_proximo:
        cutoff = datetime.now(timezone.utc) + timedelta(days=vencimiento_proximo)
        query = query.where(
            CuentaPorPagar.fecha_vencimiento <= cutoff,
            CuentaPorPagar.fecha_vencimiento >= datetime.now(timezone.utc),
            CuentaPorPagar.estado.in_(("pendiente", "parcial")),
        )

    query = query.order_by(CuentaPorPagar.fecha_vencimiento.asc()).offset(skip).limit(limit)

    result = await db.execute(query)
    cuentas = result.scalars().all()

    # Verificar vencidas y actualizar estado en memoria para la respuesta
    for cuenta in cuentas:
        _check_overdue(cuenta)

    return [CuentaPorPagarResponse.model_validate(c) for c in cuentas]


@router.get("/{cuenta_id}", response_model=CuentaPorPagarResponse)
async def get_cuenta_por_pagar(
    cuenta_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Obtener el detalle de una cuenta por pagar específica.
    """
    cuenta = await _get_cuenta_or_404(db, cuenta_id, current_user)
    _check_overdue(cuenta)
    return CuentaPorPagarResponse.model_validate(cuenta)


# ==========================================
# Registrar pago
# ==========================================

@router.post("/{cuenta_id}/pagar", response_model=CuentaPorPagarResponse)
async def registrar_pago(
    cuenta_id: str,
    payment: CuentaPorPagarPayment,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Registrar un pago (parcial o total) contra una cuenta por pagar.

    - **monto**: Monto a pagar. Si es igual al saldo pendiente, la cuenta se marca como pagada.
    - **metodo_pago**: Forma de pago (transferencia, cheque, efectivo, etc.)
    - **referencia**: Número de referencia del pago
    - **observaciones**: Notas adicionales sobre el pago
    """
    cuenta = await _get_cuenta_or_404(db, cuenta_id, current_user)

    if not cuenta.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se puede pagar una cuenta inactiva o anulada.",
        )

    if cuenta.estado == "pagada":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cuenta ya está completamente pagada.",
        )

    if payment.monto <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El monto del pago debe ser mayor a cero.",
        )

    saldo_actual = cuenta.monto_total - cuenta.monto_pagado
    if payment.monto > saldo_actual:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"El monto del pago ({payment.monto}) excede el saldo pendiente "
                f"({saldo_actual})."
            ),
        )

    cuenta.monto_pagado += payment.monto
    cuenta.monto_pendiente = cuenta.monto_total - cuenta.monto_pagado

    # Actualizar estado según el saldo restante
    if cuenta.monto_pendiente <= Decimal("0"):
        cuenta.estado = "pagada"
        cuenta.monto_pendiente = Decimal("0")
    else:
        cuenta.estado = "parcial"

    # Si hay observaciones en el pago, agregarlas
    if payment.observaciones:
        obs_prefix = f"\nPago {payment.metodo_pago or ''} ({payment.referencia or 'sin ref'}): "
        cuenta.observaciones = (cuenta.observaciones or "") + obs_prefix + payment.observaciones

    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="PAYMENT",
        entity_type="cuenta_por_pagar",
        entity_id=cuenta.id,
        description=(
            f"Pago registrado: USD {payment.monto} - Factura {cuenta.numero_factura or 'N/A'} - "
            f"Estado: {cuenta.estado}"
        ),
        ip_address=request.client.host if request.client else None,
    )

    logger.info(
        "Pago registrado en cuenta %s por USD %s (estado: %s)",
        cuenta_id,
        payment.monto,
        cuenta.estado,
    )

    return CuentaPorPagarResponse.model_validate(cuenta)


# ==========================================
# Cuentas vencidas
# ==========================================

@router.get("/vencidas", response_model=list[CuentaPorPagarResponse])
async def list_cuentas_vencidas(
    company_id: str | None = Query(None, description="ID de la empresa"),
    supplier_id: str | None = Query(None, description="ID del proveedor"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Listar cuentas por pagar vencidas (fecha_vencimiento < hoy y no pagadas).
    """
    now = datetime.now(timezone.utc)

    query = (
        select(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(
            Company.user_id == current_user.id,
            CuentaPorPagar.is_active == True,
            CuentaPorPagar.fecha_vencimiento < now,
            CuentaPorPagar.estado.in_(("pendiente", "parcial", "vencida")),
        )
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(CuentaPorPagar.company_id == company_id)
    if supplier_id:
        query = query.where(CuentaPorPagar.supplier_id == supplier_id)

    # Ordenar por las más antiguas primero (mayor urgencia)
    query = query.order_by(CuentaPorPagar.fecha_vencimiento.asc()).offset(skip).limit(limit)

    result = await db.execute(query)
    cuentas = result.scalars().all()

    # Actualizar estado a vencida si no lo está ya
    for cuenta in cuentas:
        if cuenta.estado != "vencida":
            cuenta.estado = "vencida"

    if cuentas:
        await db.flush()

    return [CuentaPorPagarResponse.model_validate(c) for c in cuentas]


# ==========================================
# Resumen
# ==========================================

@router.get("/resumen", response_model=CuentaPorPagarSummary)
async def get_resumen_cuentas(
    company_id: str | None = Query(None, description="ID de la empresa"),
    supplier_id: str | None = Query(None, description="ID del proveedor"),
    dias_proximos: int = Query(30, ge=1, le=365, description="Días para considerar 'próximo a vencer'"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Obtener resumen de cuentas por pagar: total pendiente, total vencidas,
    próximas a vencer, y desglose por estado.
    """
    now = datetime.now(timezone.utc)
    cutoff_proximo = now + timedelta(days=dias_proximos)

    base_filters = [
        Company.user_id == current_user.id,
        CuentaPorPagar.is_active == True,
        CuentaPorPagar.estado.in_(("pendiente", "parcial", "vencida")),
    ]

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        base_filters.append(CuentaPorPagar.company_id == company_id)
    if supplier_id:
        base_filters.append(CuentaPorPagar.supplier_id == supplier_id)

    # Total pendiente
    result_pendiente = await db.execute(
        select(func.sum(CuentaPorPagar.monto_pendiente))
        .select_from(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(*base_filters)
    )
    total_pendiente = result_pendiente.scalar() or Decimal("0")

    # Total vencidas
    result_vencidas = await db.execute(
        select(func.sum(CuentaPorPagar.monto_pendiente))
        .select_from(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(
            *base_filters,
            CuentaPorPagar.fecha_vencimiento < now,
        )
    )
    total_vencidas = result_vencidas.scalar() or Decimal("0")

    # Próximas a vencer
    result_proximas = await db.execute(
        select(func.sum(CuentaPorPagar.monto_pendiente))
        .select_from(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(
            *base_filters,
            CuentaPorPagar.fecha_vencimiento >= now,
            CuentaPorPagar.fecha_vencimiento <= cutoff_proximo,
        )
    )
    total_proximas = result_proximas.scalar() or Decimal("0")

    # Conteo por estado
    result_estados = await db.execute(
        select(CuentaPorPagar.estado, func.count(CuentaPorPagar.id))
        .select_from(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(*base_filters)
        .group_by(CuentaPorPagar.estado)
    )
    cuentas_por_estado = {row[0]: row[1] for row in result_estados.all()}

    # Conteo de registros
    result_count = await db.execute(
        select(func.count(CuentaPorPagar.id))
        .select_from(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .where(*base_filters)
    )
    total_cuentas = result_count.scalar() or 0

    return CuentaPorPagarSummary(
        total_pendiente=total_pendiente,
        total_vencidas=total_vencidas,
        total_proximas_a_vencer=total_proximas,
        cuentas_por_estado=cuentas_por_estado,
        total_cuentas=total_cuentas,
        dias_proximos=dias_proximos,
        generado_en=now,
    )


# ==========================================
# Exportar a Excel
# ==========================================

@router.get("/export/excel")
async def export_cuentas_excel(
    company_id: str | None = Query(None, description="ID de la empresa"),
    estado: str | None = Query(None, description="Filtrar por estado"),
    proveedor: str | None = Query(None, description="Filtrar por nombre de proveedor"),
    fecha_desde: str | None = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exportar cuentas por pagar a archivo Excel (.xlsx).

    Aplica los mismos filtros que el listado.
    """
    await _get_company_for_user(db, company_id, current_user.id) if company_id else None

    query = (
        select(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .join(Supplier, CuentaPorPagar.supplier_id == Supplier.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        query = query.where(CuentaPorPagar.company_id == company_id)
    if estado:
        query = query.where(CuentaPorPagar.estado == estado)
    if proveedor:
        query = query.where(Supplier.razon_social.ilike(f"%{proveedor}%"))
    if fecha_desde:
        try:
            desde = datetime.fromisoformat(fecha_desde).replace(tzinfo=timezone.utc)
            query = query.where(CuentaPorPagar.fecha_emision >= desde)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Formato de fecha_desde inválido. Use YYYY-MM-DD.",
            )
    if fecha_hasta:
        try:
            hasta = datetime.fromisoformat(fecha_hasta).replace(tzinfo=timezone.utc)
            hasta = hasta.replace(hour=23, minute=59, second=59)
            query = query.where(CuentaPorPagar.fecha_emision <= hasta)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Formato de fecha_hasta inválido. Use YYYY-MM-DD.",
            )

    query = query.order_by(CuentaPorPagar.fecha_vencimiento.asc())

    result = await db.execute(query)
    cuentas = result.scalars().all()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuentas por Pagar"

        headers = [
            "Nº Factura", "Proveedor", "Fecha Emisión", "Fecha Vencimiento",
            "Días Crédito", "Monto Total", "Monto Pagado", "Saldo Pendiente",
            "Estado", "Observaciones",
        ]

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        currency_format = '#,##0.00'

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Colores por estado
        estado_fills = {
            "pendiente": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            "parcial": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
            "pagada": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "vencida": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
            "anulada": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
        }

        for row_idx, cuenta in enumerate(cuentas, 2):
            dias_vencimiento = ""
            if cuenta.fecha_vencimiento:
                diff = (cuenta.fecha_vencimiento - datetime.now(timezone.utc)).days
                dias_vencimiento = diff

            values = [
                cuenta.numero_factura or "N/A",
                cuenta.supplier.razon_social if cuenta.supplier else "N/A",
                cuenta.fecha_emision.strftime("%Y-%m-%d") if cuenta.fecha_emision else "",
                cuenta.fecha_vencimiento.strftime("%Y-%m-%d") if cuenta.fecha_vencimiento else "",
                cuenta.dias_credito,
                float(cuenta.monto_total),
                float(cuenta.monto_pagado),
                float(cuenta.monto_pendiente),
                cuenta.estado,
                cuenta.observaciones or "",
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

                # Formato moneda para columnas de monto
                if col in (6, 7, 8):
                    cell.number_format = currency_format

                # Color por estado en toda la fila
                row_fill = estado_fills.get(cuenta.estado)
                if row_fill:
                    cell.fill = row_fill

        # Ajustar anchos
        column_widths = [16, 35, 14, 16, 14, 16, 16, 16, 14, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Totales
        if cuentas:
            total_row = len(cuentas) + 2
            ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
            ws.cell(row=total_row, column=6, value=sum(float(c.monto_total) for c in cuentas)).number_format = currency_format
            ws.cell(row=total_row, column=6).font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=sum(float(c.monto_pagado) for c in cuentas)).number_format = currency_format
            ws.cell(row=total_row, column=7).font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=sum(float(c.monto_pendiente) for c in cuentas)).number_format = currency_format
            ws.cell(row=total_row, column=8).font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"cuentas_por_pagar_{timestamp}.xlsx"

        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error al exportar cuentas por pagar a Excel: %s", str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo Excel: {str(e)}",
        )


# ==========================================
# Exportar a CSV
# ==========================================

@router.get("/export/csv")
async def export_cuentas_csv(
    company_id: str | None = Query(None, description="ID de la empresa"),
    estado: str | None = Query(None, description="Filtrar por estado"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exportar cuentas por pagar a archivo CSV.
    """
    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)

    query = (
        select(CuentaPorPagar)
        .join(Company, CuentaPorPagar.company_id == Company.id)
        .join(Supplier, CuentaPorPagar.supplier_id == Supplier.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        query = query.where(CuentaPorPagar.company_id == company_id)
    if estado:
        query = query.where(CuentaPorPagar.estado == estado)

    query = query.order_by(CuentaPorPagar.fecha_vencimiento.asc())

    result = await db.execute(query)
    cuentas = result.scalars().all()

    try:
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

        writer.writerow([
            "numero_factura", "proveedor", "fecha_emision", "fecha_vencimiento",
            "dias_credito", "monto_total", "monto_pagado", "monto_pendiente",
            "estado", "observaciones",
        ])

        for cuenta in cuentas:
            writer.writerow([
                cuenta.numero_factura or "",
                cuenta.supplier.razon_social if cuenta.supplier else "",
                cuenta.fecha_emision.strftime("%Y-%m-%d") if cuenta.fecha_emision else "",
                cuenta.fecha_vencimiento.strftime("%Y-%m-%d") if cuenta.fecha_vencimiento else "",
                cuenta.dias_credito,
                cuenta.monto_total,
                cuenta.monto_pagado,
                cuenta.monto_pendiente,
                cuenta.estado,
                cuenta.observaciones or "",
            ])

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=cuentas_por_pagar_{timestamp}.csv"},
        )

    except Exception as e:
        logger.error("Error al exportar cuentas por pagar a CSV: %s", str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el archivo CSV: {str(e)}",
        )


# ==========================================
# Renegociar plazos
# ==========================================

@router.post("/{cuenta_id}/renegociar", response_model=CuentaPorPagarResponse)
async def renegociar_cuenta(
    cuenta_id: str,
    renegotiation: CuentaPorPagarRenegotiation,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Renegociar los plazos de pago de una cuenta por pagar.

    Permite extender la fecha de vencimiento y/o agregar observaciones
    sobre la renegociación. Solo aplicable a cuentas pendientes o vencidas.

    - **nueva_fecha_vencimiento**: Nueva fecha de vencimiento (debe ser futura)
    - **dias_extension**: Días a extender desde la fecha actual o de vencimiento
    - **motivo**: Motivo de la renegociación (requerido)
    """
    cuenta = await _get_cuenta_or_404(db, cuenta_id, current_user)

    if not cuenta.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se puede renegociar una cuenta inactiva o anulada.",
        )

    if cuenta.estado not in ("pendiente", "parcial", "vencida"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden renegociar cuentas pendientes, parciales o vencidas.",
        )

    if cuenta.estado == "pagada":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se puede renegociar una cuenta ya pagada.",
        )

    # Calcular nueva fecha de vencimiento
    nueva_fecha = renegotiation.nueva_fecha_vencimiento
    if nueva_fecha:
        if nueva_fecha <= datetime.now(timezone.utc):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La nueva fecha de vencimiento debe ser posterior a la fecha actual.",
            )
    elif renegotiation.dias_extension and renegotiation.dias_extension > 0:
        base_fecha = cuenta.fecha_vencimiento or datetime.now(timezone.utc)
        nueva_fecha = base_fecha + timedelta(days=renegotiation.dias_extension)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Debe proporcionar una nueva fecha de vencimiento o días de extensión.",
        )

    fecha_anterior = cuenta.fecha_vencimiento
    cuenta.fecha_vencimiento = nueva_fecha

    # Si estaba vencida y se extiende, volver a pendiente/parcial según corresponda
    if cuenta.estado == "vencida":
        cuenta.estado = "parcial" if cuenta.monto_pagado > 0 else "pendiente"

    # Agregar observación de renegociación
    motivo = renegotiation.motivo or "Sin motivo especificado"
    obs_line = (
        f"\n[Renegociación {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}] "
        f"Vencimiento: {fecha_anterior.strftime('%Y-%m-%d') if fecha_anterior else 'N/A'} -> "
        f"{nueva_fecha.strftime('%Y-%m-%d')} | Motivo: {motivo}"
    )
    cuenta.observaciones = (cuenta.observaciones or "") + obs_line

    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="RENEGOTIATE",
        entity_type="cuenta_por_pagar",
        entity_id=cuenta.id,
        description=(
            f"Cuenta renegociada: Factura {cuenta.numero_factura or 'N/A'} - "
            f"Vencimiento: {fecha_anterior.strftime('%Y-%m-%d') if fecha_anterior else 'N/A'} -> "
            f"{nueva_fecha.strftime('%Y-%m-%d')} | Motivo: {motivo}"
        ),
        ip_address=request.client.host if request.client else None,
    )

    logger.info(
        "Cuenta %s renegociada: vencimiento %s -> %s",
        cuenta_id,
        fecha_anterior,
        nueva_fecha,
    )

    return CuentaPorPagarResponse.model_validate(cuenta)
