"""
ContaEC - Endpoints de Importación desde Excel/CSV
Importación masiva de productos y clientes desde archivos Excel (.xlsx) y CSV
con validación por fila y mapeo de columnas configurable
"""
import csv
import io
import logging
from decimal import Decimal, InvalidOperation
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.client import Client, TipoIdentificacion
from app.models.company import Company
from app.models.product import Product, ProductoTipo
from app.models.user import User

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/imports", tags=["Importaciones"])


# ==========================================
# Funciones auxiliares
# ==========================================

async def _get_company_for_user(
    db: AsyncSession,
    company_id: str,
    user_id: str,
) -> Company:
    """
    Obtiene una empresa verificando que pertenezca al usuario actual.

    Raises:
        HTTPException: Si la empresa no existe o no pertenece al usuario
    """
    result = await db.execute(
        select(Company).where(
            Company.id == company_id,
            Company.user_id == user_id,
            Company.is_active == True,
        )
    )
    company = result.scalars().first()
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada o no pertenece al usuario actual.",
        )
    return company


def _safe_decimal(value) -> Optional[Decimal]:
    """Convierte un valor a Decimal de forma segura"""
    if value is None:
        return None
    try:
        s = str(value).strip()
        if not s or s == "":
            return None
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _safe_str(value) -> Optional[str]:
    """Convierte un valor a string de forma segura"""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ==========================================
# Mapeo de columnas
# ==========================================

# Mapeo por defecto de columnas Excel → campos del sistema
PRODUCT_FIELD_MAPPING = {
    "codigo_principal": ["codigo_principal", "codigo", "sku", "código", "cod"],
    "codigo_auxiliar": ["codigo_auxiliar", "codigo_alt", "código auxiliar"],
    "descripcion": ["descripcion", "descripción", "nombre", "producto", "descripción"],
    "tipo": ["tipo", "tipo_producto"],
    "precio_unitario": ["precio_unitario", "precio", "valor", "costo"],
    "iva_codigo": ["iva_codigo", "código_iva", "tarifa_iva"],
    "iva_porcentaje": ["iva_porcentaje", "%_iva", "porcentaje_iva"],
    "ice_codigo": ["ice_codigo", "código_ice"],
    "ice_porcentaje": ["ice_porcentaje", "%_ice"],
    "unidad_medida": ["unidad_medida", "unidad", "medida"],
    "descuento": ["descuento", "%_descuento"],
    "codigo_barras": ["codigo_barras", "barcode", "ean", "upc", "código_barras"],
    "stock": ["stock", "cantidad", "existencia", "inventario"],
    "stock_minimo": ["stock_minimo", "mínimo", "min_stock"],
    "ubicacion": ["ubicacion", "ubicación", "posición", "bodega"],
}

CLIENT_FIELD_MAPPING = {
    "tipo_identificacion": ["tipo_identificacion", "tipo_id", "tipo_identificación"],
    "identificacion": ["identificacion", "identificación", "ruc", "cedula", "cédula", "número"],
    "razon_social": ["razon_social", "razón_social", "nombre", "cliente", "nombre_completo"],
    "direccion": ["direccion", "dirección", "dir"],
    "email": ["email", "correo", "e-mail"],
    "telefono": ["telefono", "teléfono", "tel", "telf"],
}


def _map_row_to_dict(
    row: dict,
    field_mapping: dict,
    header_map: dict | None = None,
) -> dict:
    """
    Mapea una fila de Excel/CSV a un diccionario con los campos del sistema.

    Args:
        row: Fila como diccionario {columna: valor}
        field_mapping: Mapeo de campos del sistema a posibles nombres de columna
        header_map: Mapeo explícito de columnas (opcional, sobreescribe el automático)

    Returns:
        Diccionario con los campos del sistema y sus valores
    """
    result = {}
    # Normalizar las claves del row (lowercase, sin espacios)
    normalized_row = {k.strip().lower(): v for k, v in row.items() if k}

    for system_field, possible_names in field_mapping.items():
        value = None

        # Si hay mapeo explícito, usarlo primero
        if header_map and system_field in header_map:
            col_name = header_map[system_field].strip().lower()
            value = normalized_row.get(col_name)
        else:
            # Buscar en los posibles nombres
            for name in possible_names:
                if name.lower() in normalized_row:
                    value = normalized_row[name.lower()]
                    break

        if value is not None:
            result[system_field] = value

    return result


# ==========================================
# Importación de Productos
# ==========================================

@router.post("/products/excel")
async def import_products_excel(
    company_id: str,
    file: UploadFile = File(...),
    header_map: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Importar productos desde un archivo Excel (.xlsx).

    El archivo debe tener una fila de encabezados y los datos en las filas siguientes.
    Opcionalmente se puede proveer un header_map JSON para mapear columnas.
    """
    # Verificar tipo de archivo
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser de tipo Excel (.xlsx o .xls)",
        )

    # Verificar que la empresa pertenece al usuario
    await _get_company_for_user(db, company_id, current_user.id)

    # Leer archivo Excel
    try:
        import openpyxl

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        # Obtener encabezados
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")

        # Parsear header_map si se proporciona
        import json
        parsed_map = None
        if header_map:
            try:
                parsed_map = json.loads(header_map)
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="header_map debe ser un JSON válido",
                )

        # Procesar filas
        success_count = 0
        error_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Saltar filas vacías
                continue

            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_dict(row_dict, PRODUCT_FIELD_MAPPING, parsed_map)

            # Validar campos requeridos
            required = ["codigo_principal", "descripcion", "tipo", "precio_unitario",
                        "iva_codigo", "iva_porcentaje"]
            missing = [f for f in required if f not in mapped or not mapped[f]]
            if missing:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Campos requeridos faltantes: {', '.join(missing)}",
                })
                continue

            # Validar tipo
            tipo = str(mapped["tipo"]).strip().upper()
            if tipo not in ("B", "S"):
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Tipo inválido: '{mapped['tipo']}'. Debe ser 'B' o 'S'",
                })
                continue

            # Parsear precio
            precio = _safe_decimal(mapped["precio_unitario"])
            if precio is None or precio <= 0:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Precio unitario inválido: {mapped['precio_unitario']}",
                })
                continue

            # Verificar duplicado
            codigo = str(mapped["codigo_principal"]).strip()
            existing = await db.execute(
                select(Product).where(
                    Product.company_id == company_id,
                    Product.codigo_principal == codigo,
                    Product.is_active == True,
                )
            )
            if existing.scalars().first():
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Ya existe un producto con código '{codigo}'",
                })
                continue

            # Crear producto
            try:
                product = Product(
                    company_id=company_id,
                    codigo_principal=codigo,
                    codigo_auxiliar=_safe_str(mapped.get("codigo_auxiliar")),
                    descripcion=str(mapped["descripcion"]).strip(),
                    tipo=tipo,
                    precio_unitario=precio,
                    iva_codigo=str(mapped["iva_codigo"]).strip(),
                    iva_porcentaje=_safe_decimal(mapped["iva_porcentaje"]) or Decimal("0"),
                    ice_codigo=_safe_str(mapped.get("ice_codigo")),
                    ice_porcentaje=_safe_decimal(mapped.get("ice_porcentaje")),
                    unidad_medida=_safe_str(mapped.get("unidad_medida")) or "Unidad",
                    descuento=_safe_decimal(mapped.get("descuento")) or Decimal("0"),
                    codigo_barras=_safe_str(mapped.get("codigo_barras")),
                    stock=_safe_decimal(mapped.get("stock")) or Decimal("0"),
                    stock_minimo=_safe_decimal(mapped.get("stock_minimo")) or Decimal("0"),
                    ubicacion=_safe_str(mapped.get("ubicacion")),
                )
                db.add(product)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Error al crear producto: {str(e)}",
                })

        await db.flush()
        wb.close()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}",
        )

    return {
        "message": f"Importación completada: {success_count} productos importados, {error_count} errores",
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],  # Limitar errores retornados
    }


@router.post("/products/csv")
async def import_products_csv(
    company_id: str,
    file: UploadFile = File(...),
    delimiter: str = ",",
    header_map: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Importar productos desde un archivo CSV.

    El archivo debe tener una fila de encabezados y los datos separados por el delimiter especificado.
    """
    # Verificar tipo de archivo
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser de tipo CSV (.csv)",
        )

    # Verificar que la empresa pertenece al usuario
    await _get_company_for_user(db, company_id, current_user.id)

    # Leer archivo CSV
    try:
        import json

        contents = await file.read()
        text = contents.decode("utf-8-sig")  # Soporta BOM
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        # Parsear header_map
        parsed_map = None
        if header_map:
            try:
                parsed_map = json.loads(header_map)
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="header_map debe ser un JSON válido",
                )

        # Procesar filas
        success_count = 0
        error_count = 0
        errors = []

        for row_idx, row_dict in enumerate(reader, start=2):
            if not any(row_dict.values()):
                continue

            mapped = _map_row_to_dict(row_dict, PRODUCT_FIELD_MAPPING, parsed_map)

            # Validar campos requeridos
            required = ["codigo_principal", "descripcion", "tipo", "precio_unitario",
                        "iva_codigo", "iva_porcentaje"]
            missing = [f for f in required if f not in mapped or not mapped[f]]
            if missing:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Campos requeridos faltantes: {', '.join(missing)}",
                })
                continue

            # Validar tipo
            tipo = str(mapped["tipo"]).strip().upper()
            if tipo not in ("B", "S"):
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Tipo inválido: '{mapped['tipo']}'. Debe ser 'B' o 'S'",
                })
                continue

            # Parsear precio
            precio = _safe_decimal(mapped["precio_unitario"])
            if precio is None or precio <= 0:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Precio unitario inválido: {mapped['precio_unitario']}",
                })
                continue

            # Verificar duplicado
            codigo = str(mapped["codigo_principal"]).strip()
            existing = await db.execute(
                select(Product).where(
                    Product.company_id == company_id,
                    Product.codigo_principal == codigo,
                    Product.is_active == True,
                )
            )
            if existing.scalars().first():
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Ya existe un producto con código '{codigo}'",
                })
                continue

            # Crear producto
            try:
                product = Product(
                    company_id=company_id,
                    codigo_principal=codigo,
                    codigo_auxiliar=_safe_str(mapped.get("codigo_auxiliar")),
                    descripcion=str(mapped["descripcion"]).strip(),
                    tipo=tipo,
                    precio_unitario=precio,
                    iva_codigo=str(mapped["iva_codigo"]).strip(),
                    iva_porcentaje=_safe_decimal(mapped["iva_porcentaje"]) or Decimal("0"),
                    ice_codigo=_safe_str(mapped.get("ice_codigo")),
                    ice_porcentaje=_safe_decimal(mapped.get("ice_porcentaje")),
                    unidad_medida=_safe_str(mapped.get("unidad_medida")) or "Unidad",
                    descuento=_safe_decimal(mapped.get("descuento")) or Decimal("0"),
                    codigo_barras=_safe_str(mapped.get("codigo_barras")),
                    stock=_safe_decimal(mapped.get("stock")) or Decimal("0"),
                    stock_minimo=_safe_decimal(mapped.get("stock_minimo")) or Decimal("0"),
                    ubicacion=_safe_str(mapped.get("ubicacion")),
                )
                db.add(product)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Error al crear producto: {str(e)}",
                })

        await db.flush()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar CSV: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo CSV: {str(e)}",
        )

    return {
        "message": f"Importación completada: {success_count} productos importados, {error_count} errores",
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],
    }


# ==========================================
# Importación de Clientes
# ==========================================

@router.post("/clients/excel")
async def import_clients_excel(
    company_id: str,
    file: UploadFile = File(...),
    header_map: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Importar clientes desde un archivo Excel (.xlsx).

    El archivo debe tener una fila de encabezados y los datos en las filas siguientes.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser de tipo Excel (.xlsx o .xls)",
        )

    await _get_company_for_user(db, company_id, current_user.id)

    try:
        import openpyxl
        import json

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")

        parsed_map = None
        if header_map:
            try:
                parsed_map = json.loads(header_map)
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="header_map debe ser un JSON válido",
                )

        success_count = 0
        error_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_dict(row_dict, CLIENT_FIELD_MAPPING, parsed_map)

            # Validar campos requeridos
            required = ["tipo_identificacion", "identificacion", "razon_social"]
            missing = [f for f in required if f not in mapped or not mapped[f]]
            if missing:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Campos requeridos faltantes: {', '.join(missing)}",
                })
                continue

            # Validar tipo de identificación
            tipo_id = str(mapped["tipo_identificacion"]).strip()
            valid_tipos = {"04", "05", "06", "07", "08"}
            if tipo_id not in valid_tipos:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Tipo de identificación inválido: '{tipo_id}'. Válidos: 04, 05, 06, 07, 08",
                })
                continue

            # Verificar duplicado
            ident = str(mapped["identificacion"]).strip()
            existing = await db.execute(
                select(Client).where(
                    Client.company_id == company_id,
                    Client.identificacion == ident,
                    Client.is_active == True,
                )
            )
            if existing.scalars().first():
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Ya existe un cliente con identificación '{ident}'",
                })
                continue

            # Crear cliente
            try:
                client = Client(
                    company_id=company_id,
                    tipo_identificacion=tipo_id,
                    identificacion=ident,
                    razon_social=str(mapped["razon_social"]).strip(),
                    direccion=_safe_str(mapped.get("direccion")),
                    email=_safe_str(mapped.get("email")),
                    telefono=_safe_str(mapped.get("telefono")),
                )
                db.add(client)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Error al crear cliente: {str(e)}",
                })

        await db.flush()
        wb.close()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar Excel de clientes: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}",
        )

    return {
        "message": f"Importación completada: {success_count} clientes importados, {error_count} errores",
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],
    }


@router.post("/clients/csv")
async def import_clients_csv(
    company_id: str,
    file: UploadFile = File(...),
    delimiter: str = ",",
    header_map: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Importar clientes desde un archivo CSV.
    """
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser de tipo CSV (.csv)",
        )

    await _get_company_for_user(db, company_id, current_user.id)

    try:
        import json

        contents = await file.read()
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        parsed_map = None
        if header_map:
            try:
                parsed_map = json.loads(header_map)
            except json.JSONDecodeError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="header_map debe ser un JSON válido",
                )

        success_count = 0
        error_count = 0
        errors = []

        for row_idx, row_dict in enumerate(reader, start=2):
            if not any(row_dict.values()):
                continue

            mapped = _map_row_to_dict(row_dict, CLIENT_FIELD_MAPPING, parsed_map)

            required = ["tipo_identificacion", "identificacion", "razon_social"]
            missing = [f for f in required if f not in mapped or not mapped[f]]
            if missing:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Campos requeridos faltantes: {', '.join(missing)}",
                })
                continue

            tipo_id = str(mapped["tipo_identificacion"]).strip()
            valid_tipos = {"04", "05", "06", "07", "08"}
            if tipo_id not in valid_tipos:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Tipo de identificación inválido: '{tipo_id}'",
                })
                continue

            ident = str(mapped["identificacion"]).strip()
            existing = await db.execute(
                select(Client).where(
                    Client.company_id == company_id,
                    Client.identificacion == ident,
                    Client.is_active == True,
                )
            )
            if existing.scalars().first():
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Ya existe un cliente con identificación '{ident}'",
                })
                continue

            try:
                client = Client(
                    company_id=company_id,
                    tipo_identificacion=tipo_id,
                    identificacion=ident,
                    razon_social=str(mapped["razon_social"]).strip(),
                    direccion=_safe_str(mapped.get("direccion")),
                    email=_safe_str(mapped.get("email")),
                    telefono=_safe_str(mapped.get("telefono")),
                )
                db.add(client)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append({
                    "row": row_idx,
                    "error": f"Error al crear cliente: {str(e)}",
                })

        await db.flush()

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar CSV de clientes: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo CSV: {str(e)}",
        )

    return {
        "message": f"Importación completada: {success_count} clientes importados, {error_count} errores",
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors[:50],
    }


# ==========================================
# Mapeo de columnas
# ==========================================

@router.post("/catalog/mapping")
async def get_column_mapping(
    entity: str = "products",
    current_user: User = Depends(get_current_user),
):
    """
    Obtener el mapeo de columnas disponibles para una entidad.

    Retorna los campos del sistema y los posibles nombres de columna
    que se pueden usar en los archivos de importación.
    """
    if entity == "products":
        mapping = PRODUCT_FIELD_MAPPING
    elif entity == "clients":
        mapping = CLIENT_FIELD_MAPPING
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Entidad inválida. Válidas: products, clients",
        )

    return {
        "entity": entity,
        "field_mapping": mapping,
        "required_fields": (
            ["codigo_principal", "descripcion", "tipo", "precio_unitario",
             "iva_codigo", "iva_porcentaje"]
            if entity == "products"
            else ["tipo_identificacion", "identificacion", "razon_social"]
        ),
    }
