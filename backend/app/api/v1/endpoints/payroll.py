"""
ContaEC - Endpoints de Nómina / Rol de Pago
Generación, aprobación, pago y reportes de nómina conforme
a la legislación laboral ecuatoriana
"""
import io
import logging
from datetime import datetime, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.hr_constants import (
    CENACES_RATE,
    DECIMO_CUARTO_SALARIO_BASICO,
    DECIMO_TERCERO_MONTHS,
    FONDO_RESERVA_ANIOS_MIN,
    FONDO_RESERVA_RATE,
    HORA_EXTRA_DIURNA_MULT,
    HORA_EXTRA_DOMINICAL_MULT,
    HORA_EXTRA_NOCTURNA_MULT,
    HORAS_MENSUAL_DEFAULT,
    HORAS_SEMANAL_DEFAULT,
    IESS_PERSONAL_RATE,
    IESS_PATRONAL_RATE,
    IESS_RIESGOS_RATE,
    SALARIO_BASICO_UNIFICADO_2024,
    SECAP_RATE,
    VACACIONES_DIAS_ANIO,
    VACACIONES_PROVISION_RATE,
)
from app.core.security import get_current_user
from app.models.company import Company
from app.models.employee import Employee, EstadoEmpleado
from app.models.hr_extended2 import (
    Asistencia,
    CargaFamiliar,
    EvaluacionDesempeno,
    LiquidacionLaboral,
    UtilidadesDetalle,
    UtilidadesEstado,
    UtilidadesParticipacion,
)
from app.models.payroll import EstadoRol, RolPago, RolPagoDetalle
from app.models.user import User
from app.schemas.payroll import (
    DecimoCuartoRequest,
    DecimoResponse,
    DecimoTerceroRequest,
    FondosReservaResponse,
    IESSReportResponse,
    PayrollGenerate,
    RDEPReportResponse,
    RolPagoDetalleResponse,
    RolPagoFullResponse,
    RolPagoResponse,
    VacacionesBalanceResponse,
)
from app.schemas.hr_extended2 import (
    AsistenciaCreate,
    AsistenciaResumenResponse,
    AsistenciaResponse,
    BankPaymentExportRequest,
    BankPaymentExportResponse,
    CargaFamiliarCreate,
    CargaFamiliarResponse,
    EvaluacionDesempenoCreate,
    EvaluacionDesempenoResponse,
    EvaluacionDesempenoUpdate,
    IRProgressiveCalcRequest,
    IRProgressiveCalcResponse,
    LiquidacionLaboralCreate,
    LiquidacionLaboralResponse,
    UtilidadesDetalleResponse,
    UtilidadesParticipacionCreate,
    UtilidadesParticipacionFullResponse,
    UtilidadesParticipacionResponse,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/payroll", tags=["Nómina - Rol de Pago"])


# ==========================================
# Funciones auxiliares
# ==========================================

async def _get_company_for_user(
    db: AsyncSession,
    company_id: str,
    user_id: str,
) -> Company:
    """Obtiene una empresa verificando que pertenezca al usuario actual."""
    result = await db.execute(
        select(Company).where(
            Company.id == company_id,
            Company.user_id == user_id,
            Company.is_active == True,
        )
    )
    company = result.scalars().first()
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada o no pertenece al usuario actual.",
        )
    return company


def _calcular_valor_hora(sueldo_mensual: Decimal, horas_semanal: Decimal) -> Decimal:
    """Calcula el valor por hora normal del empleado"""
    horas_mensual = horas_semanal * Decimal("4")
    if horas_mensual > 0:
        return (sueldo_mensual / horas_mensual).quantize(Decimal("0.01"))
    return Decimal("0.00")


def _calcular_detalle_empleado(
    employee: Employee,
) -> RolPagoDetalle:
    """
    Calcula todos los valores del detalle del rol de pago para un empleado.

    Conforme a la legislación ecuatoriana:
    - Aporte personal IESS: 9.45% sobre ingresos gravados
    - Aporte patronal IESS: 11.15%
    - IEE Riesgos: 0.5%
    - SECAP: 0.2%
    - CENACES: 0.1%
    - Décimo tercero: 1/12 del sueldo mensual
    - Décimo cuarto: 1/12 del salario básico unificado
    - Vacaciones: 15/360 del sueldo diario por día trabajado en el mes
    - Fondo de reserva: si tiene más de 1 año y está habilitado
    """
    sueldo_base = employee.sueldo_mensual
    horas_semanal = employee.horas_trabajo_semanal or HORAS_SEMANAL_DEFAULT
    valor_hora = _calcular_valor_hora(sueldo_base, horas_semanal)

    # Ingresos (sin horas extras por defecto - se agregan en generación)
    total_ingresos = sueldo_base

    # Descuentos empleado
    base_iess = total_ingresos  # Base imponible para IESS
    iess_personal = (base_iess * IESS_PERSONAL_RATE / Decimal("100")).quantize(Decimal("0.01")) if employee.iess_afiliado else Decimal("0.00")
    total_descuentos = iess_personal

    # Aportes empleador
    iess_patronal = (base_iess * IESS_PATRONAL_RATE / Decimal("100")).quantize(Decimal("0.01")) if employee.iess_afiliado else Decimal("0.00")
    iee_riesgos = (base_iess * IESS_RIESGOS_RATE / Decimal("100")).quantize(Decimal("0.01")) if employee.iess_afiliado else Decimal("0.00")
    secap = (base_iess * SECAP_RATE / Decimal("100")).quantize(Decimal("0.01")) if employee.iess_afiliado else Decimal("0.00")
    cenaces = (base_iess * CENACES_RATE / Decimal("100")).quantize(Decimal("0.01")) if employee.iess_afiliado else Decimal("0.00")
    total_aportes_empleador = iess_patronal + iee_riesgos + secap + cenaces

    # Décimos
    decimo_tercero = (sueldo_base / Decimal(str(DECIMO_TERCERO_MONTHS))).quantize(Decimal("0.01"))
    decimo_cuarto = (DECIMO_CUARTO_SALARIO_BASICO / Decimal(str(DECIMO_TERCERO_MONTHS))).quantize(Decimal("0.01"))

    # Vacaciones (provisión mensual)
    sueldo_diario = employee.sueldo_diario or (sueldo_base / Decimal("30.00")).quantize(Decimal("0.01"))
    vacaciones_provision = (sueldo_diario * VACACIONES_DIAS_ANIO / Decimal("12")).quantize(Decimal("0.01"))

    # Fondos de reserva
    fondos_reserva = Decimal("0.00")
    if employee.fondo_reserva and employee.anios_servicio >= FONDO_RESERVA_ANIOS_MIN:
        fondos_reserva = (sueldo_base / Decimal("12")).quantize(Decimal("0.01"))

    # Líquido a recibir
    liquido_recibir = (total_ingresos - total_descuentos).quantize(Decimal("0.01"))

    detalle = RolPagoDetalle(
        employee_id=employee.id,
        sueldo_base=sueldo_base,
        valor_horas_extras_diurnas=Decimal("0.00"),
        valor_horas_extras_nocturnas=Decimal("0.00"),
        valor_horas_extras_dominicales=Decimal("0.00"),
        comisiones=Decimal("0.00"),
        bonos=Decimal("0.00"),
        otros_ingresos=Decimal("0.00"),
        total_ingresos=total_ingresos,
        iess_personal_945=iess_personal,
        anticipo=Decimal("0.00"),
        prestamo_empresa=Decimal("0.00"),
        retencion_judicial=Decimal("0.00"),
        otros_descuentos=Decimal("0.00"),
        total_descuentos=total_descuentos,
        iess_patronal_1115=iess_patronal,
        iee_0005=iee_riesgos,
        secap_002=secap,
        cenaces_001=cenaces,
        total_aportes_empleador=total_aportes_empleador,
        decimo_tercero=decimo_tercero,
        decimo_cuarto=decimo_cuarto,
        vacaciones_provision=vacaciones_provision,
        fondos_reserva=fondos_reserva,
        liquido_recibir=liquido_recibir,
    )
    return detalle


# ==========================================
# Endpoints de Nómina
# ==========================================

@router.post("/generate", response_model=RolPagoFullResponse, status_code=status.HTTP_201_CREATED)
async def generate_payroll(
    data: PayrollGenerate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Generar el rol de pago mensual para una empresa.

    Auto-calcula todos los valores (ingresos, descuentos, aportes, décimos,
    vacaciones, fondo de reserva) para cada empleado activo de la empresa.
    Verifica que no exista ya un rol para el mismo período.
    """
    # Verificar que la empresa pertenece al usuario
    await _get_company_for_user(db, data.company_id, current_user.id)

    # Verificar que no exista un rol activo para el mismo período
    existing = await db.execute(
        select(RolPago).where(
            RolPago.company_id == data.company_id,
            RolPago.periodo_mes == data.periodo_mes,
            RolPago.periodo_anio == data.periodo_anio,
            RolPago.estado != EstadoRol.ANULADO,
            RolPago.is_active == True,
        )
    )
    if existing.scalars().first():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe un rol de pago para el período "
                   f"{data.periodo_mes}/{data.periodo_anio} en esta empresa.",
        )

    # Obtener empleados activos de la empresa
    result = await db.execute(
        select(Employee).where(
            Employee.company_id == data.company_id,
            Employee.estado == EstadoEmpleado.ACTIVO,
            Employee.is_active == True,
        )
    )
    employees = result.scalars().all()

    if not employees:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No hay empleados activos en la empresa para generar el rol de pago.",
        )

    # Crear rol de pago
    rol_pago = RolPago(
        company_id=data.company_id,
        user_id=current_user.id,
        periodo_mes=data.periodo_mes,
        periodo_anio=data.periodo_anio,
        observaciones=data.observaciones,
    )
    db.add(rol_pago)
    await db.flush()

    # Generar detalles por empleado
    total_remuneraciones = Decimal("0.00")
    total_descuentos_global = Decimal("0.00")
    total_empleador_global = Decimal("0.00")
    total_liquido_global = Decimal("0.00")

    for employee in employees:
        detalle = _calcular_detalle_empleado(employee)
        detalle.rol_pago_id = rol_pago.id
        db.add(detalle)

        total_remuneraciones += detalle.total_ingresos
        total_descuentos_global += detalle.total_descuentos
        total_empleador_global += detalle.total_aportes_empleador
        total_liquido_global += detalle.liquido_recibir

    # Actualizar totales del rol
    rol_pago.total_remuneraciones = total_remuneraciones.quantize(Decimal("0.01"))
    rol_pago.total_descuentos = total_descuentos_global.quantize(Decimal("0.01"))
    rol_pago.total_empleador = total_empleador_global.quantize(Decimal("0.01"))
    rol_pago.total_liquido = total_liquido_global.quantize(Decimal("0.01"))

    await db.flush()

    logger.info(
        f"Rol de pago generado: empresa={data.company_id}, "
        f"período={data.periodo_mes}/{data.periodo_anio}, "
        f"empleados={len(employees)}"
    )

    return RolPagoFullResponse.model_validate(rol_pago)


@router.get("", response_model=list[RolPagoResponse])
async def list_payroll(
    company_id: str | None = None,
    estado: str | None = None,
    periodo_anio: int | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Listar roles de pago de las empresas del usuario.

    Opcionalmente filtrado por empresa, estado y año.
    """
    query = (
        select(RolPago)
        .join(Company, RolPago.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(RolPago.company_id == company_id)

    if estado:
        query = query.where(RolPago.estado == estado)

    if periodo_anio:
        query = query.where(RolPago.periodo_anio == periodo_anio)

    query = query.order_by(
        RolPago.periodo_anio.desc(),
        RolPago.periodo_mes.desc(),
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    roles = result.scalars().all()

    return [RolPagoResponse.model_validate(r) for r in roles]


@router.get("/{rol_id}", response_model=RolPagoFullResponse)
async def get_payroll(
    rol_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener un rol de pago con todos sus detalles"""
    result = await db.execute(
        select(RolPago)
        .join(Company, RolPago.company_id == Company.id)
        .where(
            RolPago.id == rol_id,
            Company.user_id == current_user.id,
        )
    )
    rol = result.scalars().first()
    if not rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rol de pago no encontrado.",
        )

    return RolPagoFullResponse.model_validate(rol)


@router.put("/{rol_id}/approve", response_model=RolPagoResponse)
async def approve_payroll(
    rol_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Aprobar un rol de pago en estado borrador.

    Solo se pueden aprobar roles en estado 'borrador'.
    """
    result = await db.execute(
        select(RolPago)
        .join(Company, RolPago.company_id == Company.id)
        .where(
            RolPago.id == rol_id,
            Company.user_id == current_user.id,
        )
    )
    rol = result.scalars().first()
    if not rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rol de pago no encontrado.",
        )

    if rol.estado != EstadoRol.BORRADOR:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se pueden aprobar roles en estado 'borrador'. Estado actual: '{rol.estado}'.",
        )

    rol.estado = EstadoRol.APROBADO
    await db.flush()

    logger.info(f"Rol de pago aprobado: id={rol_id}")

    return RolPagoResponse.model_validate(rol)


@router.put("/{rol_id}/pay", response_model=RolPagoResponse)
async def pay_payroll(
    rol_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Marcar un rol de pago como pagado.

    Solo se pueden pagar roles en estado 'aprobado'.
    Actualiza los acumulados de décimos, vacaciones y fondo de reserva
    de cada empleado.
    """
    result = await db.execute(
        select(RolPago)
        .join(Company, RolPago.company_id == Company.id)
        .where(
            RolPago.id == rol_id,
            Company.user_id == current_user.id,
        )
    )
    rol = result.scalars().first()
    if not rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Rol de pago no encontrado.",
        )

    if rol.estado != EstadoRol.APROBADO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se pueden pagar roles en estado 'aprobado'. Estado actual: '{rol.estado}'.",
        )

    # Actualizar acumulados de empleados
    for detalle in rol.detalles:
        result_emp = await db.execute(
            select(Employee).where(Employee.id == detalle.employee_id)
        )
        employee = result_emp.scalars().first()
        if employee:
            employee.decimo_tercero_acumulado = (employee.decimo_tercero_acumulado + detalle.decimo_tercero).quantize(Decimal("0.01"))
            employee.decimo_cuarto_acumulado = (employee.decimo_cuarto_acumulado + detalle.decimo_cuarto).quantize(Decimal("0.01"))
            employee.fondos_reserva_acumulado = (employee.fondos_reserva_acumulado + detalle.fondos_reserva).quantize(Decimal("0.01"))

            # Actualizar vacaciones acumuladas (días)
            dias_vacaciones_mes = (Decimal(str(VACACIONES_DIAS_ANIO)) / Decimal("12")).quantize(Decimal("0.01"))
            employee.vacaciones_acumuladas_dias = (employee.vacaciones_acumuladas_dias + dias_vacaciones_mes).quantize(Decimal("0.01"))

    rol.estado = EstadoRol.PAGADO
    rol.fecha_pago = datetime.now(timezone.utc)
    await db.flush()

    logger.info(f"Rol de pago pagado: id={rol_id}")

    return RolPagoResponse.model_validate(rol)


@router.get("/employee/{employee_id}", response_model=list[RolPagoDetalleResponse])
async def get_employee_payroll_history(
    employee_id: str,
    skip: int = Query(0, ge=0),
    limit: int = Query(24, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener el historial de roles de pago de un empleado"""
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    # Obtener detalles del empleado
    result = await db.execute(
        select(RolPagoDetalle)
        .join(RolPago, RolPagoDetalle.rol_pago_id == RolPago.id)
        .where(
            RolPagoDetalle.employee_id == employee_id,
            RolPago.estado != EstadoRol.ANULADO,
        )
        .order_by(RolPago.periodo_anio.desc(), RolPago.periodo_mes.desc())
        .offset(skip)
        .limit(limit)
    )
    detalles = result.scalars().all()

    return [RolPagoDetalleResponse.model_validate(d) for d in detalles]


# ==========================================
# Décimo Tercero
# ==========================================

@router.post("/decimo-tercero", response_model=list[DecimoResponse])
async def calculate_decimo_tercero(
    data: DecimoTerceroRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular el décimo tercero para los empleados de una empresa.

    El décimo tercero equivale a un sueldo mensual por cada año trabajado,
    proporcional a los meses trabajados. Se paga en diciembre de cada año.
    """
    await _get_company_for_user(db, data.company_id, current_user.id)

    query = select(Employee).where(
        Employee.company_id == data.company_id,
        Employee.estado == EstadoEmpleado.ACTIVO,
        Employee.is_active == True,
    )

    if data.employee_id:
        query = query.where(Employee.id == data.employee_id)

    result = await db.execute(query)
    employees = result.scalars().all()

    response = []
    for emp in employees:
        # Calcular meses trabajados en el año
        fecha_inicio = emp.fecha_ingreso
        if fecha_inicio.year < data.periodo_anio:
            meses = 12
        else:
            meses = 12 - fecha_inicio.month + 1
            if emp.fecha_salida and emp.fecha_salida.year == data.periodo_anio:
                meses = emp.fecha_salida.month - fecha_inicio.month + 1
        meses = min(meses, 12)

        valor = (emp.sueldo_mensual * Decimal(str(meses)) / Decimal("12")).quantize(Decimal("0.01"))

        response.append(DecimoResponse(
            employee_id=emp.id,
            cedula=emp.cedula,
            nombre_completo=emp.nombre_completo,
            sueldo_mensual=emp.sueldo_mensual,
            meses_trabajados=meses,
            valor_decimo=valor,
        ))

    return response


# ==========================================
# Décimo Cuarto
# ==========================================

@router.post("/decimo-cuarto", response_model=list[DecimoResponse])
async def calculate_decimo_cuarto(
    data: DecimoCuartoRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular el décimo cuarto para los empleados de una empresa.

    El décimo cuarto equivale al salario básico unificado por cada año trabajado,
    proporcional a los meses trabajados. Se paga en agosto (Sierra) o marzo (Costa).
    """
    await _get_company_for_user(db, data.company_id, current_user.id)

    query = select(Employee).where(
        Employee.company_id == data.company_id,
        Employee.estado == EstadoEmpleado.ACTIVO,
        Employee.is_active == True,
    )

    if data.employee_id:
        query = query.where(Employee.id == data.employee_id)

    result = await db.execute(query)
    employees = result.scalars().all()

    response = []
    for emp in employees:
        fecha_inicio = emp.fecha_ingreso
        if fecha_inicio.year < data.periodo_anio:
            meses = 12
        else:
            meses = 12 - fecha_inicio.month + 1
            if emp.fecha_salida and emp.fecha_salida.year == data.periodo_anio:
                meses = emp.fecha_salida.month - fecha_inicio.month + 1
        meses = min(meses, 12)

        valor = (DECIMO_CUARTO_SALARIO_BASICO * Decimal(str(meses)) / Decimal("12")).quantize(Decimal("0.01"))

        response.append(DecimoResponse(
            employee_id=emp.id,
            cedula=emp.cedula,
            nombre_completo=emp.nombre_completo,
            sueldo_mensual=emp.sueldo_mensual,
            meses_trabajados=meses,
            valor_decimo=valor,
        ))

    return response


# ==========================================
# Vacaciones
# ==========================================

@router.get("/vacaciones/{employee_id}", response_model=VacacionesBalanceResponse)
async def calculate_vacation_balance(
    employee_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular el balance de vacaciones de un empleado.

    15 días de vacaciones por cada año completo de servicio.
    Se acumulan hasta un máximo de 30 días.
    """
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    anios_servicio = employee.anios_servicio
    dias_disponibles = Decimal(str(min(anios_servicio * VACACIONES_DIAS_ANIO, 30)))
    dias_acumulados = employee.vacaciones_acumuladas_dias or Decimal("0.00")

    sueldo_diario = employee.sueldo_diario or (employee.sueldo_mensual / Decimal("30.00")).quantize(Decimal("0.01"))
    valor_vacaciones = (sueldo_diario * dias_acumulados).quantize(Decimal("0.01"))

    return VacacionesBalanceResponse(
        employee_id=employee.id,
        cedula=employee.cedula,
        nombre_completo=employee.nombre_completo,
        fecha_ingreso=employee.fecha_ingreso,
        anios_servicio=anios_servicio,
        dias_disponibles=dias_disponibles,
        dias_acumulados=dias_acumulados,
        valor_vacaciones=valor_vacaciones,
    )


# ==========================================
# Fondos de Reserva
# ==========================================

@router.get("/fondos-reserva/{employee_id}", response_model=FondosReservaResponse)
async def calculate_fondos_reserva(
    employee_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular el balance de fondo de reserva de un empleado.

    El empleado tiene derecho a fondo de reserva después de 1 año de servicio.
    Equivale a un sueldo mensual por cada año completo.
    """
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    anios_servicio = employee.anios_servicio
    tiene_derecho = anios_servicio >= FONDO_RESERVA_ANIOS_MIN and employee.fondo_reserva

    return FondosReservaResponse(
        employee_id=employee.id,
        cedula=employee.cedula,
        nombre_completo=employee.nombre_completo,
        fecha_ingreso=employee.fecha_ingreso,
        anios_servicio=anios_servicio,
        tiene_derecho=tiene_derecho,
        acumulado=employee.fondos_reserva_acumulado or Decimal("0.00"),
    )


# ==========================================
# Reporte IESS
# ==========================================

@router.get("/iess/report", response_model=IESSReportResponse)
async def generate_iess_report(
    company_id: str = Query(..., description="ID de la empresa"),
    periodo_mes: int = Query(..., ge=1, le=12, description="Mes del período"),
    periodo_anio: int = Query(..., ge=2000, le=2100, description="Año del período"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Generar reporte de aportaciones al IESS para un período.

    Incluye aportes personales, patronales, IEE, SECAP y CENACES
    desglosados por empleado.
    """
    await _get_company_for_user(db, company_id, current_user.id)

    # Buscar el rol de pago del período
    result = await db.execute(
        select(RolPago).where(
            RolPago.company_id == company_id,
            RolPago.periodo_mes == periodo_mes,
            RolPago.periodo_anio == periodo_anio,
            RolPago.estado != EstadoRol.ANULADO,
            RolPago.is_active == True,
        )
    )
    rol = result.scalars().first()

    total_aporte_personal = Decimal("0.00")
    total_aporte_patronal = Decimal("0.00")
    total_iee = Decimal("0.00")
    total_secap = Decimal("0.00")
    total_cenaces = Decimal("0.00")
    detalle = []

    if rol:
        for d in rol.detalles:
            emp_result = await db.execute(
                select(Employee).where(Employee.id == d.employee_id)
            )
            emp = emp_result.scalars().first()

            total_aporte_personal += d.iess_personal_945
            total_aporte_patronal += d.iess_patronal_1115
            total_iee += d.iee_0005
            total_secap += d.secap_002
            total_cenaces += d.cenaces_001

            detalle.append({
                "employee_id": d.employee_id,
                "cedula": emp.cedula if emp else "",
                "nombre_completo": emp.nombre_completo if emp else "",
                "base_imponible": float(d.total_ingresos),
                "aporte_personal": float(d.iess_personal_945),
                "aporte_patronal": float(d.iess_patronal_1115),
                "iee_riesgos": float(d.iee_0005),
                "secap": float(d.secap_002),
                "cenaces": float(d.cenaces_001),
            })

    total_general = total_aporte_personal + total_aporte_patronal + total_iee + total_secap + total_cenaces

    return IESSReportResponse(
        company_id=company_id,
        periodo_mes=periodo_mes,
        periodo_anio=periodo_anio,
        total_empleados=len(detalle),
        total_aporte_personal=total_aporte_personal.quantize(Decimal("0.01")),
        total_aporte_patronal=total_aporte_patronal.quantize(Decimal("0.01")),
        total_aporte_iee=total_iee.quantize(Decimal("0.01")),
        total_aporte_secap=total_secap.quantize(Decimal("0.01")),
        total_aporte_cenaces=total_cenaces.quantize(Decimal("0.01")),
        total_general=total_general.quantize(Decimal("0.01")),
        detalle=detalle,
    )


# ==========================================
# Reporte RDEP
# ==========================================

@router.get("/rdep/report", response_model=RDEPReportResponse)
async def generate_rdep_report(
    company_id: str = Query(..., description="ID de la empresa"),
    periodo_mes: int = Query(..., ge=1, le=12, description="Mes del período"),
    periodo_anio: int = Query(..., ge=2000, le=2100, description="Año del período"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Generar reporte RDEP (Registro de Dependientes) para un período.

    Incluye remuneraciones, descuentos, líquido y aportes del empleador
    desglosados por empleado.
    """
    await _get_company_for_user(db, company_id, current_user.id)

    result = await db.execute(
        select(RolPago).where(
            RolPago.company_id == company_id,
            RolPago.periodo_mes == periodo_mes,
            RolPago.periodo_anio == periodo_anio,
            RolPago.estado != EstadoRol.ANULADO,
            RolPago.is_active == True,
        )
    )
    rol = result.scalars().first()

    total_remuneraciones = Decimal("0.00")
    total_descuentos = Decimal("0.00")
    total_liquido = Decimal("0.00")
    total_aportes_empleador = Decimal("0.00")
    detalle = []

    if rol:
        total_remuneraciones = rol.total_remuneraciones
        total_descuentos = rol.total_descuentos
        total_liquido = rol.total_liquido
        total_aportes_empleador = rol.total_empleador

        for d in rol.detalles:
            emp_result = await db.execute(
                select(Employee).where(Employee.id == d.employee_id)
            )
            emp = emp_result.scalars().first()

            detalle.append({
                "employee_id": d.employee_id,
                "cedula": emp.cedula if emp else "",
                "nombre_completo": emp.nombre_completo if emp else "",
                "cargo": emp.cargo if emp else "",
                "total_ingresos": float(d.total_ingresos),
                "total_descuentos": float(d.total_descuentos),
                "liquido_recibir": float(d.liquido_recibir),
                "total_aportes_empleador": float(d.total_aportes_empleador),
            })

    return RDEPReportResponse(
        company_id=company_id,
        periodo_mes=periodo_mes,
        periodo_anio=periodo_anio,
        total_empleados=len(detalle),
        total_remuneraciones=total_remuneraciones.quantize(Decimal("0.01")),
        total_descuentos=total_descuentos.quantize(Decimal("0.01")),
        total_liquido=total_liquido.quantize(Decimal("0.01")),
        total_aportes_empleador=total_aportes_empleador.quantize(Decimal("0.01")),
        detalle=detalle,
    )


# ==========================================
# Exportar a Excel
# ==========================================

@router.get("/export/excel")
async def export_payroll_excel(
    company_id: str = Query(..., description="ID de la empresa"),
    periodo_mes: int = Query(..., ge=1, le=12, description="Mes del período"),
    periodo_anio: int = Query(..., ge=2000, le=2100, description="Año del período"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exportar el rol de pago a archivo Excel (.xlsx).

    Genera un archivo Excel con todos los detalles del rol de pago
    incluyendo ingresos, descuentos, aportes y líquido por empleado.
    """
    await _get_company_for_user(db, company_id, current_user.id)

    result = await db.execute(
        select(RolPago).where(
            RolPago.company_id == company_id,
            RolPago.periodo_mes == periodo_mes,
            RolPago.periodo_anio == periodo_anio,
            RolPago.estado != EstadoRol.ANULADO,
            RolPago.is_active == True,
        )
    )
    rol = result.scalars().first()

    if not rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontró rol de pago para el período especificado.",
        )

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Librería openpyxl no instalada. Instale con: pip install openpyxl",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rol_{periodo_mes}_{periodo_anio}"

    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Encabezado
    ws.merge_cells("A1:T1")
    ws["A1"] = f"ROL DE PAGO - {periodo_mes:02d}/{periodo_anio}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers de columna
    headers = [
        "Cédula", "Apellidos y Nombres", "Cargo", "Departamento",
        "Sueldo Base", "H.E. Diurnas", "V.H.E. Diurnas",
        "H.E. Nocturnas", "V.H.E. Nocturnas", "H.E. Dominicales", "V.H.E. Dominicales",
        "Comisiones", "Bonos", "Otros Ingresos", "Total Ingresos",
        "IESS Personal 9.45%", "Anticipo", "Préstamo", "Retención Judicial", "Otros Descuentos",
        "Total Descuentos", "IESS Patronal 11.15%", "IEE 0.5%", "SECAP 0.2%", "CENACES 0.1%",
        "Total Aportes Empleador", "Décimo Tercero", "Décimo Cuarto",
        "Vacaciones", "Fondos Reserva", "Líquido a Recibir",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Datos por empleado
    row = 4
    for d in rol.detalles:
        emp_result = await db.execute(
            select(Employee).where(Employee.id == d.employee_id)
        )
        emp = emp_result.scalars().first()
        if not emp:
            continue

        data_row = [
            emp.cedula, emp.nombre_completo, emp.cargo, emp.departamento or "",
            float(d.sueldo_base), float(d.horas_extras_diurnas), float(d.valor_horas_extras_diurnas),
            float(d.horas_extras_nocturnas), float(d.valor_horas_extras_nocturnas),
            float(d.horas_extras_dominicales), float(d.valor_horas_extras_dominicales),
            float(d.comisiones), float(d.bonos), float(d.otros_ingresos), float(d.total_ingresos),
            float(d.iess_personal_945), float(d.anticipo), float(d.prestamo_empresa),
            float(d.retencion_judicial), float(d.otros_descuentos), float(d.total_descuentos),
            float(d.iess_patronal_1115), float(d.iee_0005), float(d.secap_002), float(d.cenaces_001),
            float(d.total_aportes_empleador), float(d.decimo_tercero), float(d.decimo_cuarto),
            float(d.vacaciones_provision), float(d.fondos_reserva), float(d.liquido_recibir),
        ]

        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

        row += 1

    # Fila de totales
    row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=row, column=15, value=float(rol.total_remuneraciones)).font = Font(bold=True)
    ws.cell(row=row, column=21, value=float(rol.total_descuentos)).font = Font(bold=True)
    ws.cell(row=row, column=26, value=float(rol.total_empleador)).font = Font(bold=True)
    ws.cell(row=row, column=31, value=float(rol.total_liquido)).font = Font(bold=True)

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"rol_pago_{periodo_mes:02d}_{periodo_anio}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==========================================
# Cargas Familiares
# ==========================================

@router.post("/cargas-familiares", response_model=CargaFamiliarResponse, status_code=status.HTTP_201_CREATED)
async def create_carga_familiar(
    data: CargaFamiliarCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Registrar una carga familiar para un empleado.

    Permite agregar hijos, cónyuge u otros dependientes del empleado
    para efectos de impuesto a la renta y beneficios sociales.
    """
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == data.employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    carga = CargaFamiliar(
        employee_id=data.employee_id,
        nombres=data.nombres,
        apellidos=data.apellidos,
        parentesco=data.parentesco,
        fecha_nacimiento=data.fecha_nacimiento,
        identificacion=data.identificacion,
        discapacidad=data.discapacidad,
        es_estudiante=data.es_estudiante,
    )
    db.add(carga)
    await db.flush()

    logger.info(f"Carga familiar creada: employee_id={data.employee_id}, parentesco={data.parentesco}")

    return CargaFamiliarResponse.model_validate(carga)


@router.get("/cargas-familiares/{employee_id}", response_model=list[CargaFamiliarResponse])
async def list_cargas_familiares(
    employee_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Listar las cargas familiares de un empleado"""
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    result = await db.execute(
        select(CargaFamiliar).where(
            CargaFamiliar.employee_id == employee_id,
            CargaFamiliar.is_active == True,
        ).order_by(CargaFamiliar.created_at.desc())
    )
    cargas = result.scalars().all()

    return [CargaFamiliarResponse.model_validate(c) for c in cargas]


@router.delete("/cargas-familiares/{carga_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_carga_familiar(
    carga_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Eliminar (desactivar) una carga familiar"""
    result = await db.execute(
        select(CargaFamiliar)
        .join(Employee, CargaFamiliar.employee_id == Employee.id)
        .join(Company, Employee.company_id == Company.id)
        .where(
            CargaFamiliar.id == carga_id,
            Company.user_id == current_user.id,
        )
    )
    carga = result.scalars().first()
    if not carga:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Carga familiar no encontrada.",
        )

    carga.is_active = False
    await db.flush()

    logger.info(f"Carga familiar eliminada: id={carga_id}")


# ==========================================
# Evaluaciones de Desempeño
# ==========================================

@router.post("/evaluaciones", response_model=EvaluacionDesempenoResponse, status_code=status.HTTP_201_CREATED)
async def create_evaluacion(
    data: EvaluacionDesempenoCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Crear una evaluación de desempeño para un empleado.

    Permite registrar el puntaje, objetivos, fortalezas, áreas de mejora
    y plan de acción para un período específico.
    """
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == data.employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    evaluacion = EvaluacionDesempeno(
        employee_id=data.employee_id,
        evaluador_id=data.evaluador_id or current_user.id,
        periodo=data.periodo,
        puntaje=data.puntaje,
        objetivos=data.objetivos,
        comentarios=data.comentarios,
        fortalezas=data.fortalezas,
        areas_mejora=data.areas_mejora,
        plan_accion=data.plan_accion,
        estado=data.estado,
        fecha_evaluacion=data.fecha_evaluacion,
    )
    db.add(evaluacion)
    await db.flush()

    logger.info(f"Evaluación creada: employee_id={data.employee_id}, periodo={data.periodo}")

    return EvaluacionDesempenoResponse.model_validate(evaluacion)


@router.get("/evaluaciones", response_model=list[EvaluacionDesempenoResponse])
async def list_evaluaciones(
    employee_id: str | None = None,
    periodo: str | None = None,
    estado: str | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Listar evaluaciones de desempeño.

    Filtrado opcional por empleado, período y estado.
    """
    query = (
        select(EvaluacionDesempeno)
        .join(Employee, EvaluacionDesempeno.employee_id == Employee.id)
        .join(Company, Employee.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if employee_id:
        query = query.where(EvaluacionDesempeno.employee_id == employee_id)
    if periodo:
        query = query.where(EvaluacionDesempeno.periodo == periodo)
    if estado:
        query = query.where(EvaluacionDesempeno.estado == estado)

    query = query.order_by(
        EvaluacionDesempeno.created_at.desc()
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    evaluaciones = result.scalars().all()

    return [EvaluacionDesempenoResponse.model_validate(e) for e in evaluaciones]


@router.put("/evaluaciones/{evaluacion_id}", response_model=EvaluacionDesempenoResponse)
async def update_evaluacion(
    evaluacion_id: str,
    data: EvaluacionDesempenoUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Actualizar una evaluación de desempeño"""
    result = await db.execute(
        select(EvaluacionDesempeno)
        .join(Employee, EvaluacionDesempeno.employee_id == Employee.id)
        .join(Company, Employee.company_id == Company.id)
        .where(
            EvaluacionDesempeno.id == evaluacion_id,
            Company.user_id == current_user.id,
        )
    )
    evaluacion = result.scalars().first()
    if not evaluacion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evaluación no encontrada.",
        )

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(evaluacion, field, value)

    await db.flush()

    logger.info(f"Evaluación actualizada: id={evaluacion_id}")

    return EvaluacionDesempenoResponse.model_validate(evaluacion)


# ==========================================
# Asistencia
# ==========================================

@router.post("/asistencia", response_model=AsistenciaResponse, status_code=status.HTTP_201_CREATED)
async def create_asistencia(
    data: AsistenciaCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Registrar la asistencia de un empleado.

    Registra la hora de entrada, salida, horas trabajadas y tipo de día.
    Compatible con datos de dispositivos biométricos.
    """
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == data.employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    asistencia = Asistencia(
        employee_id=data.employee_id,
        fecha=data.fecha,
        hora_entrada=data.hora_entrada,
        hora_salida=data.hora_salida,
        horas_trabajadas=data.horas_trabajadas,
        horas_extras=data.horas_extras,
        tipo=data.tipo,
        dispositivo_entrada=data.dispositivo_entrada,
        dispositivo_salida=data.dispositivo_salida,
        observacion=data.observacion,
    )
    db.add(asistencia)
    await db.flush()

    logger.info(f"Asistencia registrada: employee_id={data.employee_id}, fecha={data.fecha}")

    return AsistenciaResponse.model_validate(asistencia)


@router.get("/asistencia", response_model=list[AsistenciaResponse])
async def list_asistencia(
    employee_id: str | None = None,
    fecha_desde: datetime | None = None,
    fecha_hasta: datetime | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Listar registros de asistencia.

    Filtrado opcional por empleado y rango de fechas.
    """
    query = (
        select(Asistencia)
        .join(Employee, Asistencia.employee_id == Employee.id)
        .join(Company, Employee.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if employee_id:
        query = query.where(Asistencia.employee_id == employee_id)
    if fecha_desde:
        query = query.where(Asistencia.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.where(Asistencia.fecha <= fecha_hasta)

    query = query.order_by(
        Asistencia.fecha.desc()
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    asistencias = result.scalars().all()

    return [AsistenciaResponse.model_validate(a) for a in asistencias]


@router.get("/asistencia/resumen", response_model=list[AsistenciaResumenResponse])
async def get_asistencia_resumen(
    employee_id: str | None = None,
    anio: int = Query(..., ge=2000, le=2100, description="Año"),
    mes: int = Query(..., ge=1, le=12, description="Mes"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Obtener resumen mensual de asistencia.

    Agrega horas trabajadas, horas extras y conteo de días por tipo
    (normal, descanso, festivo, vacaciones, permiso, enfermedad).
    """
    query = (
        select(Asistencia)
        .join(Employee, Asistencia.employee_id == Employee.id)
        .join(Company, Employee.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if employee_id:
        query = query.where(Asistencia.employee_id == employee_id)

    # Filter by month/year
    from sqlalchemy import extract
    query = query.where(
        extract("year", Asistencia.fecha) == anio,
        extract("month", Asistencia.fecha) == mes,
    )

    result = await db.execute(query)
    asistencias = result.scalars().all()

    # Group by employee
    resumen_por_empleado: dict[str, dict] = {}
    for a in asistencias:
        if a.employee_id not in resumen_por_empleado:
            resumen_por_empleado[a.employee_id] = {
                "dias_trabajados": 0,
                "dias_descanso": 0,
                "dias_festivo": 0,
                "dias_vacacion": 0,
                "dias_permiso": 0,
                "dias_enfermedad": 0,
                "total_horas_trabajadas": Decimal("0.00"),
                "total_horas_extras": Decimal("0.00"),
            }

        r = resumen_por_empleado[a.employee_id]
        if a.tipo == "normal":
            r["dias_trabajados"] += 1
        elif a.tipo == "descanso":
            r["dias_descanso"] += 1
        elif a.tipo == "festivo":
            r["dias_festivo"] += 1
        elif a.tipo == "vacacion":
            r["dias_vacacion"] += 1
        elif a.tipo == "permiso":
            r["dias_permiso"] += 1
        elif a.tipo == "enfermedad":
            r["dias_enfermedad"] += 1

        r["total_horas_trabajadas"] += a.horas_trabajadas
        r["total_horas_extras"] += a.horas_extras

    response = []
    for emp_id, resumen in resumen_por_empleado.items():
        response.append(AsistenciaResumenResponse(
            employee_id=emp_id,
            anio=anio,
            mes=mes,
            dias_trabajados=resumen["dias_trabajados"],
            dias_descanso=resumen["dias_descanso"],
            dias_festivo=resumen["dias_festivo"],
            dias_vacacion=resumen["dias_vacacion"],
            dias_permiso=resumen["dias_permiso"],
            dias_enfermedad=resumen["dias_enfermedad"],
            total_horas_trabajadas=resumen["total_horas_trabajadas"].quantize(Decimal("0.01")),
            total_horas_extras=resumen["total_horas_extras"].quantize(Decimal("0.01")),
        ))

    return response


@router.post("/asistencia/import", response_model=list[AsistenciaResponse], status_code=status.HTTP_201_CREATED)
async def import_asistencia(
    employee_id: str = Query(..., description="ID del empleado"),
    registros: list[AsistenciaCreate] = ...,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Importar registros de asistencia desde CSV/Excel de dispositivo biométrico.

    Permite carga masiva de registros de asistencia.
    """
    # Verificar que el empleado pertenece a una empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == employee_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado.",
        )

    created = []
    for reg in registros:
        asistencia = Asistencia(
            employee_id=employee_id,
            fecha=reg.fecha,
            hora_entrada=reg.hora_entrada,
            hora_salida=reg.hora_salida,
            horas_trabajadas=reg.horas_trabajadas,
            horas_extras=reg.horas_extras,
            tipo=reg.tipo,
            dispositivo_entrada=reg.dispositivo_entrada,
            dispositivo_salida=reg.dispositivo_salida,
            observacion=reg.observacion,
        )
        db.add(asistencia)
        created.append(asistencia)

    await db.flush()

    logger.info(f"Asistencia importada: employee_id={employee_id}, registros={len(created)}")

    return [AsistenciaResponse.model_validate(a) for a in created]


# ==========================================
# Liquidación Laboral
# ==========================================

@router.post("/liquidacion", response_model=LiquidacionLaboralResponse, status_code=status.HTTP_201_CREATED)
async def calculate_liquidacion(
    data: LiquidacionLaboralCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular la liquidación laboral de un empleado.

    Calcula automáticamente: sueldo pendiente, décimo tercero/cuarto,
    vacaciones, fondo de reserva, bono de desahucio e indemnización
    conforme al Código del Trabajo ecuatoriano.
    """
    # Verificar que el empleado pertenece a la empresa del usuario
    result = await db.execute(
        select(Employee)
        .join(Company, Employee.company_id == Company.id)
        .where(
            Employee.id == data.employee_id,
            Employee.company_id == data.company_id,
            Company.user_id == current_user.id,
        )
    )
    employee = result.scalars().first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empleado no encontrado en la empresa indicada.",
        )

    fecha_inicio = employee.fecha_ingreso
    fecha_fin = data.fecha_fin

    # Calcular años y meses de servicio
    delta = fecha_fin - fecha_inicio
    dias_totales = delta.days
    anios_servicio = dias_totales // 365
    meses_servicio = dias_totales // 30

    sueldo_diario = (employee.sueldo_mensual / Decimal("30.00")).quantize(Decimal("0.01"))

    # Sueldo pendiente (días del mes actual no pagados)
    dias_mes_pendiente = fecha_fin.day if fecha_fin else 0
    sueldo_pendiente = (sueldo_diario * Decimal(str(dias_mes_pendiente))).quantize(Decimal("0.01"))

    # Décimo tercero pendiente
    decimo_tercero_pendiente = employee.decimo_tercero_acumulado or Decimal("0.00")

    # Décimo cuarto pendiente
    decimo_cuarto_pendiente = employee.decimo_cuarto_acumulado or Decimal("0.00")

    # Vacaciones pendientes (monetizadas)
    dias_vacaciones = employee.vacaciones_acumuladas_dias or Decimal("0.00")
    vacaciones_pendientes = (sueldo_diario * dias_vacaciones).quantize(Decimal("0.01"))

    # Fondo de reserva pendiente
    fondo_reserva_pendiente = employee.fondos_reserva_acumulado or Decimal("0.00")

    # Bono de desahucio (1 sueldo por cada año, solo en despido/renuncia voluntaria)
    bono_desahucio = Decimal("0.00")
    if data.tipo in ("despido", "renuncia_voluntaria") and anios_servicio > 0:
        bono_desahucio = (employee.sueldo_mensual * Decimal(str(anios_servicio))).quantize(Decimal("0.01"))
        # Máximo 25 sueldos
        maximo_bono = (employee.sueldo_mensual * Decimal("25")).quantize(Decimal("0.01"))
        if bono_desahucio > maximo_bono:
            bono_desahucio = maximo_bono

    # Indemnización por despido intempestivo (3 sueldos + 1 por cada año)
    indemnizacion = Decimal("0.00")
    if data.tipo == "despido":
        indemnizacion = (employee.sueldo_mensual * Decimal("3")).quantize(Decimal("0.01"))
        if anios_servicio > 0:
            indemnizacion += (employee.sueldo_mensual * Decimal(str(anios_servicio))).quantize(Decimal("0.01"))
            # Máximo 25 sueldos adicionales
            maximo_indemnizacion = (employee.sueldo_mensual * Decimal("25")).quantize(Decimal("0.01"))
            if indemnizacion > maximo_indemnizacion + (employee.sueldo_mensual * Decimal("3")):
                indemnizacion = (employee.sueldo_mensual * Decimal("28")).quantize(Decimal("0.01"))

    # Total ingresos
    total_ingresos = (
        sueldo_pendiente
        + decimo_tercero_pendiente
        + decimo_cuarto_pendiente
        + vacaciones_pendientes
        + fondo_reserva_pendiente
        + bono_desahucio
        + indemnizacion
        + data.otros_ingresos
    ).quantize(Decimal("0.01"))

    # Total descuentos
    total_descuentos = (
        data.iess_pendiente
        + data.anticipos_pendientes
        + data.prestamos_pendientes
        + data.otros_descuentos
    ).quantize(Decimal("0.01"))

    # Total liquidación
    total_liquidacion = (total_ingresos - total_descuentos).quantize(Decimal("0.01"))

    liquidacion = LiquidacionLaboral(
        employee_id=data.employee_id,
        company_id=data.company_id,
        tipo=data.tipo,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        sueldo_pendiente=sueldo_pendiente,
        decimo_tercero_pendiente=decimo_tercero_pendiente,
        decimo_cuarto_pendiente=decimo_cuarto_pendiente,
        vacaciones_pendientes=vacaciones_pendientes,
        fondo_reserva_pendiente=fondo_reserva_pendiente,
        bono_desahucio=bono_desahucio,
        indemnizacion=indemnizacion,
        otros_ingresos=data.otros_ingresos,
        total_ingresos=total_ingresos,
        iess_pendiente=data.iess_pendiente,
        anticipos_pendientes=data.anticipos_pendientes,
        prestamos_pendientes=data.prestamos_pendientes,
        otros_descuentos=data.otros_descuentos,
        total_descuentos=total_descuentos,
        total_liquidacion=total_liquidacion,
        observaciones=data.observaciones,
    )
    db.add(liquidacion)
    await db.flush()

    logger.info(
        f"Liquidación calculada: employee_id={data.employee_id}, "
        f"tipo={data.tipo}, total={total_liquidacion}"
    )

    return LiquidacionLaboralResponse.model_validate(liquidacion)


@router.get("/liquidacion", response_model=list[LiquidacionLaboralResponse])
async def list_liquidaciones(
    company_id: str | None = None,
    employee_id: str | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Listar liquidaciones laborales filtradas por empresa y/o empleado"""
    query = (
        select(LiquidacionLaboral)
        .join(Company, LiquidacionLaboral.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(LiquidacionLaboral.company_id == company_id)

    if employee_id:
        query = query.where(LiquidacionLaboral.employee_id == employee_id)

    query = query.order_by(
        LiquidacionLaboral.created_at.desc()
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    liquidaciones = result.scalars().all()

    return [LiquidacionLaboralResponse.model_validate(l) for l in liquidaciones]


@router.put("/liquidacion/{liquidacion_id}/approve", response_model=LiquidacionLaboralResponse)
async def approve_liquidacion(
    liquidacion_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Aprobar una liquidación laboral.

    Solo se pueden aprobar liquidaciones en estado 'borrador'.
    """
    result = await db.execute(
        select(LiquidacionLaboral)
        .join(Company, LiquidacionLaboral.company_id == Company.id)
        .where(
            LiquidacionLaboral.id == liquidacion_id,
            Company.user_id == current_user.id,
        )
    )
    liquidacion = result.scalars().first()
    if not liquidacion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Liquidación no encontrada.",
        )

    if liquidacion.estado != "borrador":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se pueden aprobar liquidaciones en estado 'borrador'. Estado actual: '{liquidacion.estado}'.",
        )

    liquidacion.estado = "aprobada"
    liquidacion.aprobado_por = current_user.id
    liquidacion.fecha_aprobacion = datetime.now(timezone.utc)
    await db.flush()

    logger.info(f"Liquidación aprobada: id={liquidacion_id}")

    return LiquidacionLaboralResponse.model_validate(liquidacion)


@router.put("/liquidacion/{liquidacion_id}/pay", response_model=LiquidacionLaboralResponse)
async def pay_liquidacion(
    liquidacion_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Marcar una liquidación laboral como pagada.

    Solo se pueden pagar liquidaciones en estado 'aprobada'.
    """
    result = await db.execute(
        select(LiquidacionLaboral)
        .join(Company, LiquidacionLaboral.company_id == Company.id)
        .where(
            LiquidacionLaboral.id == liquidacion_id,
            Company.user_id == current_user.id,
        )
    )
    liquidacion = result.scalars().first()
    if not liquidacion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Liquidación no encontrada.",
        )

    if liquidacion.estado != "aprobada":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se pueden pagar liquidaciones en estado 'aprobada'. Estado actual: '{liquidacion.estado}'.",
        )

    liquidacion.estado = "pagada"
    await db.flush()

    # Actualizar estado del empleado a cese
    emp_result = await db.execute(
        select(Employee).where(Employee.id == liquidacion.employee_id)
    )
    employee = emp_result.scalars().first()
    if employee and employee.estado != "cese":
        employee.estado = EstadoEmpleado.CESE
        employee.fecha_salida = liquidacion.fecha_fin

    logger.info(f"Liquidación pagada: id={liquidacion_id}")

    return LiquidacionLaboralResponse.model_validate(liquidacion)


# ==========================================
# Utilidades / Participación de Utilidades
# ==========================================

@router.post("/utilidades", response_model=UtilidadesParticipacionFullResponse, status_code=status.HTTP_201_CREATED)
async def calculate_utilidades(
    data: UtilidadesParticipacionCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular la distribución de utilidades (15% para trabajadores).

    Conforme al Art. 97 de la Ley de Compañías y el Código del Trabajo:
    - 10% se distribuye por igual entre todos los trabajadores
    - 5% se distribuye en proporción a las cargas familiares
    Se calcula en base a días trabajados y sueldo acumulado.
    """
    await _get_company_for_user(db, data.company_id, current_user.id)

    # Verificar que no exista ya una distribución para el año
    existing = await db.execute(
        select(UtilidadesParticipacion).where(
            UtilidadesParticipacion.company_id == data.company_id,
            UtilidadesParticipacion.anio == data.anio,
        )
    )
    if existing.scalars().first():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ya existe una distribución de utilidades para el año {data.anio} en esta empresa.",
        )

    # Obtener empleados activos
    result = await db.execute(
        select(Employee).where(
            Employee.company_id == data.company_id,
            Employee.estado == EstadoEmpleado.ACTIVO,
            Employee.is_active == True,
        )
    )
    employees = result.scalars().all()

    if not employees:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No hay empleados activos para calcular la distribución de utilidades.",
        )

    # Calcular monto a distribuir
    monto_distribuir = (
        data.total_utilidades * data.porcentaje_trabajadores / Decimal("100")
    ).quantize(Decimal("0.01"))

    # Calcular total de días trabajados y sueldos acumulados
    total_dias = 0
    total_sueldos = Decimal("0.00")
    empleados_data = []

    for emp in employees:
        # Calcular días trabajados en el año
        fecha_inicio = emp.fecha_ingreso
        if fecha_inicio.year < data.anio:
            dias = 365
        else:
            dias = (datetime(data.anio, 12, 31, tzinfo=timezone.utc) - fecha_inicio).days
            if emp.fecha_salida and emp.fecha_salida.year == data.anio:
                dias = (emp.fecha_salida - fecha_inicio).days
        dias = max(dias, 0)
        dias = min(dias, 365)

        sueldo_acumulado = (emp.sueldo_mensual * Decimal(str(dias)) / Decimal("30")).quantize(Decimal("0.01"))

        total_dias += dias
        total_sueldos += sueldo_acumulado

        empleados_data.append({
            "employee_id": emp.id,
            "dias_trabajados": dias,
            "sueldo_acumulado": sueldo_acumulado,
        })

    # Crear registro principal
    utilidad = UtilidadesParticipacion(
        company_id=data.company_id,
        anio=data.anio,
        total_utilidades=data.total_utilidades,
        porcentaje_trabajadores=data.porcentaje_trabajadores,
        numero_trabajadores=len(employees),
        monto_distribuir=monto_distribuir,
    )
    db.add(utilidad)
    await db.flush()

    # Calcular detalle por empleado (proporcional a días trabajados)
    for emp_data in empleados_data:
        if total_dias > 0:
            porcentaje = (Decimal(str(emp_data["dias_trabajados"])) / Decimal(str(total_dias)) * Decimal("100")).quantize(Decimal("0.0001"))
        else:
            porcentaje = Decimal("0.0000")

        monto_asignado = (monto_distribuir * Decimal(str(emp_data["dias_trabajados"])) / Decimal(str(total_dias))).quantize(Decimal("0.01")) if total_dias > 0 else Decimal("0.00")

        detalle = UtilidadesDetalle(
            utilidad_id=utilidad.id,
            employee_id=emp_data["employee_id"],
            dias_trabajados=emp_data["dias_trabajados"],
            sueldo_acumulado=emp_data["sueldo_acumulado"],
            porcentaje_participacion=porcentaje,
            monto_asignado=monto_asignado,
        )
        db.add(detalle)

    await db.flush()

    logger.info(
        f"Utilidades calculadas: company_id={data.company_id}, "
        f"anio={data.anio}, monto_distribuir={monto_distribuir}"
    )

    # Re-fetch to get detalles
    result = await db.execute(
        select(UtilidadesParticipacion).where(UtilidadesParticipacion.id == utilidad.id)
    )
    utilidad = result.scalars().first()

    return UtilidadesParticipacionFullResponse.model_validate(utilidad)


@router.get("/utilidades", response_model=list[UtilidadesParticipacionResponse])
async def list_utilidades(
    company_id: str | None = None,
    anio: int | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Listar distribuciones de utilidades filtradas por empresa y/o año"""
    query = (
        select(UtilidadesParticipacion)
        .join(Company, UtilidadesParticipacion.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(UtilidadesParticipacion.company_id == company_id)

    if anio:
        query = query.where(UtilidadesParticipacion.anio == anio)

    query = query.order_by(
        UtilidadesParticipacion.anio.desc()
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    utilidades = result.scalars().all()

    return [UtilidadesParticipacionResponse.model_validate(u) for u in utilidades]


@router.put("/utilidades/{utilidad_id}/approve", response_model=UtilidadesParticipacionResponse)
async def approve_utilidades(
    utilidad_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Aprobar una distribución de utilidades.

    Solo se pueden aprobar distribuciones en estado 'borrador'.
    """
    result = await db.execute(
        select(UtilidadesParticipacion)
        .join(Company, UtilidadesParticipacion.company_id == Company.id)
        .where(
            UtilidadesParticipacion.id == utilidad_id,
            Company.user_id == current_user.id,
        )
    )
    utilidad = result.scalars().first()
    if not utilidad:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Distribución de utilidades no encontrada.",
        )

    if utilidad.estado != UtilidadesEstado.BORRADOR:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se pueden aprobar distribuciones en estado 'borrador'. Estado actual: '{utilidad.estado}'.",
        )

    utilidad.estado = UtilidadesEstado.APROBADA
    await db.flush()

    logger.info(f"Utilidades aprobadas: id={utilidad_id}")

    return UtilidadesParticipacionResponse.model_validate(utilidad)


@router.get("/utilidades/{utilidad_id}/detalle", response_model=UtilidadesParticipacionFullResponse)
async def get_utilidades_detalle(
    utilidad_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener el detalle de distribución de utilidades por empleado"""
    result = await db.execute(
        select(UtilidadesParticipacion)
        .join(Company, UtilidadesParticipacion.company_id == Company.id)
        .where(
            UtilidadesParticipacion.id == utilidad_id,
            Company.user_id == current_user.id,
        )
    )
    utilidad = result.scalars().first()
    if not utilidad:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Distribución de utilidades no encontrada.",
        )

    return UtilidadesParticipacionFullResponse.model_validate(utilidad)


# ==========================================
# Impuesto a la Renta (IR) Progresivo
# ==========================================

# Tabla de Impuesto a la Renta Ecuador 2024
IR_TABLE_2024 = [
    # (fracción_básica, exceso_hasta, tasa%, impuesto_fracción_básica)
    (Decimal("0.00"), Decimal("11310.00"), Decimal("0.00"), Decimal("0.00")),
    (Decimal("11310.00"), Decimal("14410.00"), Decimal("5.00"), Decimal("0.00")),
    (Decimal("14410.00"), Decimal("18010.00"), Decimal("10.00"), Decimal("155.00")),
    (Decimal("18010.00"), Decimal("21630.00"), Decimal("12.00"), Decimal("515.00")),
    (Decimal("21630.00"), Decimal("30310.00"), Decimal("15.00"), Decimal("949.40")),
    (Decimal("30310.00"), Decimal("40440.00"), Decimal("20.00"), Decimal("2252.40")),
    (Decimal("40440.00"), Decimal("53470.00"), Decimal("25.00"), Decimal("4278.40")),
    (Decimal("53470.00"), Decimal("71940.00"), Decimal("30.00"), Decimal("7535.90")),
    (Decimal("71940.00"), Decimal("95260.00"), Decimal("32.00"), Decimal("13082.90")),
    (Decimal("95260.00"), Decimal("127040.00"), Decimal("33.00"), Decimal("20529.50")),
    (Decimal("127040.00"), Decimal("169380.00"), Decimal("34.00"), Decimal("31002.30")),
    (Decimal("169380.00"), Decimal("225810.00"), Decimal("35.00"), Decimal("45398.30")),
    (Decimal("225810.00"), None, Decimal("35.00"), Decimal("65112.85")),
]


@router.post("/calcular-ir", response_model=IRProgressiveCalcResponse)
async def calcular_impuesto_renta(
    data: IRProgressiveCalcRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Calcular el Impuesto a la Renta progresivo según tabla Ecuador 2024.

    Fracción básica exenta: $11,310.00
    Tasas progresivas desde 0% hasta 35%.
    """
    ingreso_gravable = data.ingreso_gravable

    # Encontrar el tramo correspondiente
    fraccion_basica = Decimal("0.00")
    tasa = Decimal("0.00")
    impuesto_fraccion = Decimal("0.00")
    exceso = Decimal("0.00")

    for i, (fb_desde, fb_hasta, t, imp_fb) in enumerate(IR_TABLE_2024):
        if fb_hasta is None or ingreso_gravable <= fb_hasta:
            fraccion_basica = fb_desde
            tasa = t
            impuesto_fraccion = imp_fb
            exceso = (ingreso_gravable - fb_desde).quantize(Decimal("0.01"))
            break

    # Calcular impuesto
    impuesto_exceso = (exceso * tasa / Decimal("100")).quantize(Decimal("0.01"))
    impuesto_total = (impuesto_fraccion + impuesto_exceso).quantize(Decimal("0.01"))

    # Base imponible es el ingreso menos la fracción básica exenta
    base_imponible = max(ingreso_gravable - Decimal("11310.00"), Decimal("0.00")).quantize(Decimal("0.01"))

    return IRProgressiveCalcResponse(
        base_imponible=base_imponible,
        impuesto=impuesto_total,
        fraccion_basica=fraccion_basica,
        exceso=exceso,
        tasa=tasa,
        impuesto_fraccion=impuesto_fraccion,
        fecha=data.periodo,
    )


# ==========================================
# Exportación de Pago Bancario
# ==========================================

@router.get("/export-bank-payment", response_model=BankPaymentExportResponse)
async def export_bank_payment(
    company_id: str = Query(..., description="ID de la empresa"),
    periodo: str = Query(..., description="Período de pago (ej: 2024-01)"),
    formato: str = Query(default="xlsx", description="Formato: xlsx o csv"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Exportar archivo de pago bancario para el período indicado.

    Genera un archivo Excel o CSV con los datos de pago de cada empleado
    (banco, tipo de cuenta, número de cuenta, monto a pagar) para ser
    importado directamente por el banco.
    """
    await _get_company_for_user(db, company_id, current_user.id)

    # Parsear período
    try:
        anio, mes = int(periodo.split("-")[0]), int(periodo.split("-")[1])
    except (ValueError, IndexError):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de período inválido. Use YYYY-MM (ej: 2024-01).",
        )

    # Buscar rol de pago del período
    result = await db.execute(
        select(RolPago).where(
            RolPago.company_id == company_id,
            RolPago.periodo_mes == mes,
            RolPago.periodo_anio == anio,
            RolPago.estado != EstadoRol.ANULADO,
            RolPago.is_active == True,
        )
    )
    rol = result.scalars().first()

    if not rol:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontró rol de pago para el período {periodo}.",
        )

    # Preparar datos
    pagos = []
    for detalle in rol.detalles:
        emp_result = await db.execute(
            select(Employee).where(Employee.id == detalle.employee_id)
        )
        emp = emp_result.scalars().first()
        if emp and emp.banco and emp.numero_cuenta:
            pagos.append({
                "cedula": emp.cedula,
                "nombres": emp.nombre_completo,
                "banco": emp.banco,
                "tipo_cuenta": emp.tipo_cuenta or "",
                "numero_cuenta": emp.numero_cuenta,
                "monto": float(detalle.liquido_recibir),
            })

    if formato == "csv":
        import csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["cedula", "nombres", "banco", "tipo_cuenta", "numero_cuenta", "monto"])
        writer.writeheader()
        writer.writerows(pagos)
        content = output.getvalue()
        filename = f"pago_bancario_{periodo}.csv"
    else:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Librería openpyxl no instalada.",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Pago_{periodo}"

        headers = ["Cédula", "Nombres", "Banco", "Tipo Cuenta", "Número Cuenta", "Monto"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, pago in enumerate(pagos, 2):
            ws.cell(row=row_idx, column=1, value=pago["cedula"])
            ws.cell(row=row_idx, column=2, value=pago["nombres"])
            ws.cell(row=row_idx, column=3, value=pago["banco"])
            ws.cell(row=row_idx, column=4, value=pago["tipo_cuenta"])
            ws.cell(row=row_idx, column=5, value=pago["numero_cuenta"])
            ws.cell(row=row_idx, column=6, value=pago["monto"])

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        filename = f"pago_bancario_{periodo}.xlsx"

    # Encode to base64
    import base64
    if isinstance(content, str):
        content_bytes = content.encode("utf-8")
    else:
        content_bytes = content

    content_b64 = base64.b64encode(content_bytes).decode("utf-8")

    return BankPaymentExportResponse(
        filename=filename,
        content=content_b64,
    )
