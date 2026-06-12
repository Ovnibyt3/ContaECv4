"""
ContaEC - Endpoints de Presupuestos y Control Presupuestario
CRUD de presupuestos anuales, cuentas, ejecución mensual, alertas y comparativos
"""
import logging
from datetime import datetime, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_action
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.budget import (
    PresupuestoAlerta,
    PresupuestoAnual,
    PresupuestoCuenta,
    PresupuestoEjecucionMensual,
    PresupuestoEstado,
    TipoAlertaPresupuesto,
    TipoCuenta,
)
from app.models.company import Company
from app.models.user import User
from app.schemas.budget import (
    BudgetAlertItem,
    BudgetAlertsResponse,
    BudgetExecutionDetailResponse,
    BudgetExecutionMonthDetail,
    BudgetExportResponse,
    ComparativoGeneralResponse,
    ComparativoPresupuestario,
    EjecucionMensualCreate,
    EjecucionMensualUpdate,
    PresupuestoAnualCreate,
    PresupuestoAnualResponse,
    PresupuestoAnualUpdate,
    PresupuestoAlertaResponse,
    PresupuestoCuentaCreate,
    PresupuestoCuentaResponse,
    PresupuestoCuentaUpdate,
    PresupuestoEjecucionMensualResponse,
    PresupuestoStatsResponse,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/budgets", tags=["Presupuestos"])


# ==========================================
# Funciones auxiliares
# ==========================================

async def _get_company_for_user(
    db: AsyncSession,
    company_id: str,
    user_id: str,
) -> Company:
    """Obtiene una empresa verificando que pertenezca al usuario actual"""
    result = await db.execute(
        select(Company).where(
            Company.id == company_id,
            Company.user_id == user_id,
            Company.is_active == True,
        )
    )
    company = result.scalars().first()
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada o no pertenece al usuario actual.",
        )
    return company


async def _check_alerts(
    db: AsyncSession,
    cuenta: PresupuestoCuenta,
    company_id: str,
) -> list[PresupuestoAlerta]:
    """
    Verifica umbrales de ejecución y genera alertas automáticas.
    Umbrales: 50%, 75%, 90% y sobregiro (>100%).
    Solo genera alerta si no existe una igual (misma cuenta + tipo) sin resolver.
    """
    alertas_creadas = []
    if cuenta.monto_anual == 0:
        return alertas_creadas

    pct = cuenta.porcentaje_ejecucion

    thresholds = [
        (TipoAlertaPresupuesto.CINCUENTA_PORCIENTO, Decimal("50")),
        (TipoAlertaPresupuesto.SETENTA_CINCO_PORCIENTO, Decimal("75")),
        (TipoAlertaPresupuesto.NOVENTA_PORCIENTO, Decimal("90")),
    ]

    for tipo_alerta, threshold in thresholds:
        if pct >= threshold:
            # Check if un leída/unresolved alert of same type exists
            existing = await db.execute(
                select(PresupuestoAlerta).where(
                    PresupuestoAlerta.presupuesto_cuenta_id == cuenta.id,
                    PresupuestoAlerta.tipo_alerta == tipo_alerta.value,
                    PresupuestoAlerta.is_resuelta == False,
                )
            )
            if not existing.scalars().first():
                alerta = PresupuestoAlerta(
                    company_id=company_id,
                    presupuesto_cuenta_id=cuenta.id,
                    tipo_alerta=tipo_alerta.value,
                    mensaje=(
                        f"La cuenta {cuenta.cuenta_codigo} - {cuenta.cuenta_nombre} "
                        f"ha alcanzado el {threshold}% de ejecución presupuestaria "
                        f"({pct:.2f}% ejecutado)."
                    ),
                    monto_presupuestado=cuenta.monto_anual,
                    monto_ejecutado=cuenta.monto_ejecutado,
                    monto_sobregiro=Decimal("0"),
                    porcentaje_ejecucion=pct,
                )
                db.add(alerta)
                alertas_creadas.append(alerta)

    # Check for sobregiro (ejecutado > presupuestado)
    if cuenta.monto_ejecutado > cuenta.monto_anual:
        existing_sobregiro = await db.execute(
            select(PresupuestoAlerta).where(
                PresupuestoAlerta.presupuesto_cuenta_id == cuenta.id,
                PresupuestoAlerta.tipo_alerta == TipoAlertaPresupuesto.SOBREGIRO.value,
                PresupuestoAlerta.is_resuelta == False,
            )
        )
        if not existing_sobregiro.scalars().first():
            monto_sobregiro = cuenta.monto_ejecutado - cuenta.monto_anual
            alerta = PresupuestoAlerta(
                company_id=company_id,
                presupuesto_cuenta_id=cuenta.id,
                tipo_alerta=TipoAlertaPresupuesto.SOBREGIRO.value,
                mensaje=(
                    f"SOBREGIRO en la cuenta {cuenta.cuenta_codigo} - {cuenta.cuenta_nombre}. "
                    f"Presupuestado: ${cuenta.monto_anual:.2f}, "
                    f"Ejecutado: ${cuenta.monto_ejecutado:.2f}, "
                    f"Sobregiro: ${monto_sobregiro:.2f}."
                ),
                monto_presupuestado=cuenta.monto_anual,
                monto_ejecutado=cuenta.monto_ejecutado,
                monto_sobregiro=monto_sobregiro,
                porcentaje_ejecucion=pct,
            )
            db.add(alerta)
            alertas_creadas.append(alerta)

    return alertas_creadas


async def _recalculate_cuenta(cuenta: PresupuestoCuenta) -> None:
    """Recalcula totales de una cuenta presupuestaria a partir de ejecuciones mensuales"""
    monto_ejecutado = Decimal("0")
    for ejecucion in cuenta.ejecuciones_mensuales:
        monto_ejecutado += ejecucion.monto_ejecutado

    cuenta.monto_ejecutado = monto_ejecutado
    cuenta.monto_disponible = cuenta.monto_anual - monto_ejecutado

    if cuenta.monto_anual > 0:
        cuenta.porcentaje_ejecucion = (
            (monto_ejecutado / cuenta.monto_anual * Decimal("100"))
            .quantize(Decimal("0.01"))
        )
    else:
        cuenta.porcentaje_ejecucion = Decimal("0")


async def _recalculate_presupuesto(presupuesto: PresupuestoAnual) -> None:
    """Recalcula totales del presupuesto anual a partir de las cuentas"""
    total_ingresos_presupuestado = Decimal("0")
    total_egresos_presupuestado = Decimal("0")
    total_ingresos_ejecutado = Decimal("0")
    total_egresos_ejecutado = Decimal("0")

    for cuenta in presupuesto.cuentas:
        if not cuenta.is_active:
            continue
        if cuenta.cuenta_tipo == TipoCuenta.INGRESO.value:
            total_ingresos_presupuestado += cuenta.monto_anual
            total_ingresos_ejecutado += cuenta.monto_ejecutado
        elif cuenta.cuenta_tipo == TipoCuenta.EGRESO.value:
            total_egresos_presupuestado += cuenta.monto_anual
            total_egresos_ejecutado += cuenta.monto_ejecutado

    presupuesto.total_ingresos_presupuestado = total_ingresos_presupuestado
    presupuesto.total_egresos_presupuestado = total_egresos_presupuestado
    presupuesto.total_ingresos_ejecutado = total_ingresos_ejecutado
    presupuesto.total_egresos_ejecutado = total_egresos_ejecutado


# ==========================================
# 1. CRUD Presupuestos Anuales
# ==========================================

@router.post("", response_model=PresupuestoAnualResponse, status_code=status.HTTP_201_CREATED)
async def create_presupuesto(
    data: PresupuestoAnualCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Crear un nuevo presupuesto anual con cuentas"""
    await _get_company_for_user(db, data.company_id, current_user.id)

    # Validate cuenta_tipo values
    valid_tipos = {TipoCuenta.INGRESO.value, TipoCuenta.EGRESO.value}
    for cuenta_data in data.cuentas:
        if cuenta_data.cuenta_tipo not in valid_tipos:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tipo de cuenta inválido: {cuenta_data.cuenta_tipo}. Debe ser 'ingreso' o 'egreso'.",
            )

    # Create presupuesto
    presupuesto = PresupuestoAnual(
        company_id=data.company_id,
        user_id=current_user.id,
        anio=data.anio,
        nombre=data.nombre,
        descripcion=data.descripcion,
        estado=PresupuestoEstado.BORRADOR.value,
    )

    # Create cuentas
    cuentas = []
    for cuenta_data in data.cuentas:
        cuenta = PresupuestoCuenta(
            cuenta_codigo=cuenta_data.cuenta_codigo,
            cuenta_nombre=cuenta_data.cuenta_nombre,
            cuenta_tipo=cuenta_data.cuenta_tipo,
            monto_anual=cuenta_data.monto_anual,
            monto_ejecutado=Decimal("0"),
            monto_disponible=cuenta_data.monto_anual,
            porcentaje_ejecucion=Decimal("0"),
        )

        # Distribución mensual
        if cuenta_data.distribucion_mensual:
            if len(cuenta_data.distribucion_mensual) != 12:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="La distribución mensual debe tener exactamente 12 valores.",
                )
            distribucion = cuenta_data.distribucion_mensual
        else:
            # Distribute evenly
            mensual = (cuenta_data.monto_anual / Decimal("12")).quantize(Decimal("0.01"))
            distribucion = [mensual] * 12

        ejecuciones = []
        for mes_num, monto_mensual in enumerate(distribucion, start=1):
            ejecucion = PresupuestoEjecucionMensual(
                mes=mes_num,
                monto_presupuestado=monto_mensual,
                monto_ejecutado=Decimal("0"),
                monto_disponible=monto_mensual,
                porcentaje_ejecucion=Decimal("0"),
            )
            ejecuciones.append(ejecucion)

        cuenta.ejecuciones_mensuales = ejecuciones
        cuentas.append(cuenta)

    presupuesto.cuentas = cuentas
    db.add(presupuesto)
    await db.flush()

    # Recalculate totals
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        entity_type="presupuesto_anual",
        entity_id=presupuesto.id,
        description=f"Presupuesto anual creado: {presupuesto.nombre} ({presupuesto.anio})",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoAnualResponse.model_validate(presupuesto)


@router.get("", response_model=list[PresupuestoAnualResponse])
async def list_presupuestos(
    company_id: str | None = None,
    anio: int | None = None,
    estado: str | None = None,
    is_active: bool | None = True,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Listar presupuestos anuales con filtros"""
    query = (
        select(PresupuestoAnual)
        .join(Company, PresupuestoAnual.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(PresupuestoAnual.company_id == company_id)
    if anio is not None:
        query = query.where(PresupuestoAnual.anio == anio)
    if estado:
        query = query.where(PresupuestoAnual.estado == estado)
    if is_active is not None:
        query = query.where(PresupuestoAnual.is_active == is_active)

    query = query.order_by(PresupuestoAnual.anio.desc(), PresupuestoAnual.created_at.desc()).offset(skip).limit(limit)

    result = await db.execute(query)
    presupuestos = result.scalars().all()
    return [PresupuestoAnualResponse.model_validate(p) for p in presupuestos]


@router.get("/stats", response_model=PresupuestoStatsResponse)
async def get_presupuesto_stats(
    company_id: str = Query(..., description="ID de la empresa"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener estadísticas generales de presupuestos"""
    await _get_company_for_user(db, company_id, current_user.id)

    # Total presupuestos
    total_result = await db.execute(
        select(func.count(PresupuestoAnual.id)).where(
            PresupuestoAnual.company_id == company_id,
            PresupuestoAnual.is_active == True,
        )
    )
    total = total_result.scalar() or 0

    # By estado
    borrador_result = await db.execute(
        select(func.count(PresupuestoAnual.id)).where(
            PresupuestoAnual.company_id == company_id,
            PresupuestoAnual.estado == PresupuestoEstado.BORRADOR.value,
            PresupuestoAnual.is_active == True,
        )
    )
    borrador = borrador_result.scalar() or 0

    aprobados_result = await db.execute(
        select(func.count(PresupuestoAnual.id)).where(
            PresupuestoAnual.company_id == company_id,
            PresupuestoAnual.estado == PresupuestoEstado.APROBADO.value,
            PresupuestoAnual.is_active == True,
        )
    )
    aprobados = aprobados_result.scalar() or 0

    cerrados_result = await db.execute(
        select(func.count(PresupuestoAnual.id)).where(
            PresupuestoAnual.company_id == company_id,
            PresupuestoAnual.estado == PresupuestoEstado.CERRADO.value,
            PresupuestoAnual.is_active == True,
        )
    )
    cerrados = cerrados_result.scalar() or 0

    # Cuentas con alertas (sin resolver)
    cuentas_con_alerta_result = await db.execute(
        select(func.count(func.distinct(PresupuestoAlerta.presupuesto_cuenta_id))).where(
            PresupuestoAlerta.company_id == company_id,
            PresupuestoAlerta.is_resuelta == False,
        )
    )
    cuentas_con_alerta = cuentas_con_alerta_result.scalar() or 0

    # Total sobregiros
    sobregiros_result = await db.execute(
        select(func.count(PresupuestoAlerta.id)).where(
            PresupuestoAlerta.company_id == company_id,
            PresupuestoAlerta.tipo_alerta == TipoAlertaPresupuesto.SOBREGIRO.value,
            PresupuestoAlerta.is_resuelta == False,
        )
    )
    sobregiros = sobregiros_result.scalar() or 0

    return PresupuestoStatsResponse(
        total_presupuestos=total,
        presupuestos_borrador=borrador,
        presupuestos_aprobados=aprobados,
        presupuestos_cerrados=cerrados,
        total_cuentas_con_alerta=cuentas_con_alerta,
        total_sobregiros=sobregiros,
    )


@router.get("/comparativo", response_model=ComparativoGeneralResponse)
async def get_comparativo_general(
    company_id: str = Query(..., description="ID de la empresa"),
    anio: int = Query(..., description="Año fiscal"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Comparativo presupuestado vs ejecutado para un año y empresa"""
    await _get_company_for_user(db, company_id, current_user.id)

    # Get presupuesto for the year
    result = await db.execute(
        select(PresupuestoAnual).where(
            PresupuestoAnual.company_id == company_id,
            PresupuestoAnual.anio == anio,
            PresupuestoAnual.is_active == True,
        )
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontró presupuesto activo para el año {anio}.",
        )

    # Build comparativo por cuenta
    cuentas_comparativo = []
    total_ingresos_presupuestado = Decimal("0")
    total_ingresos_ejecutado = Decimal("0")
    total_egresos_presupuestado = Decimal("0")
    total_egresos_ejecutado = Decimal("0")

    for cuenta in presupuesto.cuentas:
        if not cuenta.is_active:
            continue

        variacion = cuenta.monto_anual - cuenta.monto_ejecutado
        variacion_pct = (
            (variacion / cuenta.monto_anual * Decimal("100")).quantize(Decimal("0.01"))
            if cuenta.monto_anual > 0
            else Decimal("0")
        )

        # Check for active alert
        alerta_tipo = None
        for alerta in cuenta.alertas:
            if not alerta.is_resuelta:
                alerta_tipo = alerta.tipo_alerta
                break

        cuentas_comparativo.append(ComparativoPresupuestario(
            cuenta_codigo=cuenta.cuenta_codigo,
            cuenta_nombre=cuenta.cuenta_nombre,
            cuenta_tipo=cuenta.cuenta_tipo,
            monto_presupuestado=cuenta.monto_anual,
            monto_ejecutado=cuenta.monto_ejecutado,
            monto_disponible=cuenta.monto_disponible,
            porcentaje_ejecucion=cuenta.porcentaje_ejecucion,
            variacion=variacion,
            variacion_porcentaje=variacion_pct,
            alerta_tipo=alerta_tipo,
        ))

        if cuenta.cuenta_tipo == TipoCuenta.INGRESO.value:
            total_ingresos_presupuestado += cuenta.monto_anual
            total_ingresos_ejecutado += cuenta.monto_ejecutado
        else:
            total_egresos_presupuestado += cuenta.monto_anual
            total_egresos_ejecutado += cuenta.monto_ejecutado

    resultado_presupuestario = total_ingresos_presupuestado - total_egresos_presupuestado
    resultado_real = total_ingresos_ejecutado - total_egresos_ejecutado

    return ComparativoGeneralResponse(
        anio=anio,
        total_ingresos_presupuestado=total_ingresos_presupuestado,
        total_ingresos_ejecutado=total_ingresos_ejecutado,
        total_egresos_presupuestado=total_egresos_presupuestado,
        total_egresos_ejecutado=total_egresos_ejecutado,
        resultado_presupuestario=resultado_presupuestario,
        resultado_real=resultado_real,
        cuentas=cuentas_comparativo,
    )


@router.get("/alertas/summary")
async def get_alertas_summary(
    company_id: str = Query(..., description="ID de la empresa"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener resumen de alertas presupuestarias"""
    await _get_company_for_user(db, company_id, current_user.id)

    # Count by tipo
    tipo_counts = {}
    for tipo in TipoAlertaPresupuesto:
        result = await db.execute(
            select(func.count(PresupuestoAlerta.id)).where(
                PresupuestoAlerta.company_id == company_id,
                PresupuestoAlerta.tipo_alerta == tipo.value,
                PresupuestoAlerta.is_resuelta == False,
            )
        )
        tipo_counts[tipo.value] = result.scalar() or 0

    # Total no leídas
    no_leidas_result = await db.execute(
        select(func.count(PresupuestoAlerta.id)).where(
            PresupuestoAlerta.company_id == company_id,
            PresupuestoAlerta.is_leida == False,
            PresupuestoAlerta.is_resuelta == False,
        )
    )
    no_leidas = no_leidas_result.scalar() or 0

    # Total sin resolver
    sin_resolver_result = await db.execute(
        select(func.count(PresupuestoAlerta.id)).where(
            PresupuestoAlerta.company_id == company_id,
            PresupuestoAlerta.is_resuelta == False,
        )
    )
    sin_resolver = sin_resolver_result.scalar() or 0

    return {
        "total_no_leidas": no_leidas,
        "total_sin_resolver": sin_resolver,
        "por_tipo": tipo_counts,
    }


@router.get("/alertas", response_model=list[PresupuestoAlertaResponse])
async def list_alertas(
    company_id: str | None = None,
    tipo: str | None = None,
    is_leida: bool | None = None,
    is_resuelta: bool | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Listar alertas presupuestarias con filtros"""
    query = (
        select(PresupuestoAlerta)
        .join(Company, PresupuestoAlerta.company_id == Company.id)
        .where(Company.user_id == current_user.id)
    )

    if company_id:
        await _get_company_for_user(db, company_id, current_user.id)
        query = query.where(PresupuestoAlerta.company_id == company_id)
    if tipo:
        query = query.where(PresupuestoAlerta.tipo_alerta == tipo)
    if is_leida is not None:
        query = query.where(PresupuestoAlerta.is_leida == is_leida)
    if is_resuelta is not None:
        query = query.where(PresupuestoAlerta.is_resuelta == is_resuelta)

    query = query.order_by(PresupuestoAlerta.created_at.desc()).offset(skip).limit(limit)

    result = await db.execute(query)
    alertas = result.scalars().all()
    return [PresupuestoAlertaResponse.model_validate(a) for a in alertas]


@router.get("/{presupuesto_id}", response_model=PresupuestoAnualResponse)
async def get_presupuesto(
    presupuesto_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener un presupuesto anual con cuentas y ejecuciones"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)
    return PresupuestoAnualResponse.model_validate(presupuesto)


@router.get("/{presupuesto_id}/comparativo", response_model=ComparativoGeneralResponse)
async def get_comparativo_presupuesto(
    presupuesto_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Comparativo detallado para un presupuesto específico"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    # Build comparativo
    cuentas_comparativo = []
    for cuenta in presupuesto.cuentas:
        if not cuenta.is_active:
            continue

        variacion = cuenta.monto_anual - cuenta.monto_ejecutado
        variacion_pct = (
            (variacion / cuenta.monto_anual * Decimal("100")).quantize(Decimal("0.01"))
            if cuenta.monto_anual > 0
            else Decimal("0")
        )

        alerta_tipo = None
        for alerta in cuenta.alertas:
            if not alerta.is_resuelta:
                alerta_tipo = alerta.tipo_alerta
                break

        cuentas_comparativo.append(ComparativoPresupuestario(
            cuenta_codigo=cuenta.cuenta_codigo,
            cuenta_nombre=cuenta.cuenta_nombre,
            cuenta_tipo=cuenta.cuenta_tipo,
            monto_presupuestado=cuenta.monto_anual,
            monto_ejecutado=cuenta.monto_ejecutado,
            monto_disponible=cuenta.monto_disponible,
            porcentaje_ejecucion=cuenta.porcentaje_ejecucion,
            variacion=variacion,
            variacion_porcentaje=variacion_pct,
            alerta_tipo=alerta_tipo,
        ))

    resultado_presupuestario = presupuesto.total_ingresos_presupuestado - presupuesto.total_egresos_presupuestado
    resultado_real = presupuesto.total_ingresos_ejecutado - presupuesto.total_egresos_ejecutado

    return ComparativoGeneralResponse(
        anio=presupuesto.anio,
        total_ingresos_presupuestado=presupuesto.total_ingresos_presupuestado,
        total_ingresos_ejecutado=presupuesto.total_ingresos_ejecutado,
        total_egresos_presupuestado=presupuesto.total_egresos_presupuestado,
        total_egresos_ejecutado=presupuesto.total_egresos_ejecutado,
        resultado_presupuestario=resultado_presupuestario,
        resultado_real=resultado_real,
        cuentas=cuentas_comparativo,
    )


@router.put("/{presupuesto_id}", response_model=PresupuestoAnualResponse)
async def update_presupuesto(
    presupuesto_id: str,
    data: PresupuestoAnualUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Actualizar un presupuesto anual (solo en estado borrador)"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.BORRADOR.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden modificar presupuestos en estado borrador.",
        )

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(presupuesto, field, value)

    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="UPDATE",
        entity_type="presupuesto_anual",
        entity_id=presupuesto.id,
        description=f"Presupuesto actualizado: {presupuesto.nombre}",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoAnualResponse.model_validate(presupuesto)


@router.delete("/{presupuesto_id}")
async def delete_presupuesto(
    presupuesto_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Eliminar un presupuesto anual (solo en estado borrador)"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.BORRADOR.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden eliminar presupuestos en estado borrador.",
        )

    presupuesto.is_active = False
    presupuesto.estado = PresupuestoEstado.ANULADO.value
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="DELETE",
        entity_type="presupuesto_anual",
        entity_id=presupuesto.id,
        description=f"Presupuesto anulado: {presupuesto.nombre}",
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Presupuesto anulado exitosamente."}


# ==========================================
# 6-7. Aprobar / Cerrar presupuesto
# ==========================================

@router.put("/{presupuesto_id}/approve", response_model=PresupuestoAnualResponse)
async def approve_presupuesto(
    presupuesto_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Aprobar un presupuesto (borrador → aprobado)"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.BORRADOR.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se puede aprobar un presupuesto en estado '{presupuesto.estado}'. Solo se pueden aprobar presupuestos en borrador.",
        )

    presupuesto.estado = PresupuestoEstado.APROBADO.value
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="APPROVE",
        entity_type="presupuesto_anual",
        entity_id=presupuesto.id,
        description=f"Presupuesto aprobado: {presupuesto.nombre} ({presupuesto.anio})",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoAnualResponse.model_validate(presupuesto)


@router.put("/{presupuesto_id}/close", response_model=PresupuestoAnualResponse)
async def close_presupuesto(
    presupuesto_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cerrar un presupuesto (aprobado → cerrado)"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.APROBADO.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se puede cerrar un presupuesto en estado '{presupuesto.estado}'. Solo se pueden cerrar presupuestos aprobados.",
        )

    presupuesto.estado = PresupuestoEstado.CERRADO.value
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CLOSE",
        entity_type="presupuesto_anual",
        entity_id=presupuesto.id,
        description=f"Presupuesto cerrado: {presupuesto.nombre} ({presupuesto.anio})",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoAnualResponse.model_validate(presupuesto)


# ==========================================
# 8-10. CRUD Cuentas Presupuestarias
# ==========================================

@router.post("/{presupuesto_id}/cuentas", response_model=PresupuestoCuentaResponse, status_code=status.HTTP_201_CREATED)
async def add_cuenta_to_presupuesto(
    presupuesto_id: str,
    data: PresupuestoCuentaCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Agregar una cuenta a un presupuesto (solo en borrador)"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.BORRADOR.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden agregar cuentas a presupuestos en borrador.",
        )

    # Validate cuenta_tipo
    valid_tipos = {TipoCuenta.INGRESO.value, TipoCuenta.EGRESO.value}
    if data.cuenta_tipo not in valid_tipos:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tipo de cuenta inválido: {data.cuenta_tipo}.",
        )

    cuenta = PresupuestoCuenta(
        presupuesto_id=presupuesto_id,
        cuenta_codigo=data.cuenta_codigo,
        cuenta_nombre=data.cuenta_nombre,
        cuenta_tipo=data.cuenta_tipo,
        monto_anual=data.monto_anual,
        monto_ejecutado=Decimal("0"),
        monto_disponible=data.monto_anual,
        porcentaje_ejecucion=Decimal("0"),
    )

    # Distribución mensual
    if data.distribucion_mensual:
        if len(data.distribucion_mensual) != 12:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La distribución mensual debe tener exactamente 12 valores.",
            )
        distribucion = data.distribucion_mensual
    else:
        mensual = (data.monto_anual / Decimal("12")).quantize(Decimal("0.01"))
        distribucion = [mensual] * 12

    ejecuciones = []
    for mes_num, monto_mensual in enumerate(distribucion, start=1):
        ejecucion = PresupuestoEjecucionMensual(
            mes=mes_num,
            monto_presupuestado=monto_mensual,
            monto_ejecutado=Decimal("0"),
            monto_disponible=monto_mensual,
            porcentaje_ejecucion=Decimal("0"),
        )
        ejecuciones.append(ejecucion)

    cuenta.ejecuciones_mensuales = ejecuciones
    db.add(cuenta)
    await db.flush()

    # Recalculate presupuesto totals
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        entity_type="presupuesto_cuenta",
        entity_id=cuenta.id,
        description=f"Cuenta presupuestaria agregada: {cuenta.cuenta_codigo} - {cuenta.cuenta_nombre}",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoCuentaResponse.model_validate(cuenta)


@router.put("/cuentas/{cuenta_id}", response_model=PresupuestoCuentaResponse)
async def update_cuenta(
    cuenta_id: str,
    data: PresupuestoCuentaUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Actualizar monto de una cuenta presupuestaria (solo en borrador)"""
    result = await db.execute(
        select(PresupuestoCuenta).where(PresupuestoCuenta.id == cuenta_id)
    )
    cuenta = result.scalars().first()
    if not cuenta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cuenta presupuestaria no encontrada.",
        )

    # Get presupuesto to verify ownership and state
    presup_result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == cuenta.presupuesto_id)
    )
    presupuesto = presup_result.scalars().first()
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.BORRADOR.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden modificar cuentas de presupuestos en borrador.",
        )

    update_data = data.model_dump(exclude_unset=True)

    if "monto_anual" in update_data and update_data["monto_anual"] is not None:
        new_monto = update_data["monto_anual"]
        old_monto = cuenta.monto_anual
        cuenta.monto_anual = new_monto
        cuenta.monto_disponible = new_monto - cuenta.monto_ejecutado

        # Redistribute monthly budgets proportionally
        if old_monto > 0 and cuenta.ejecuciones_mensuales:
            ratio = new_monto / old_monto
            for ejecucion in cuenta.ejecuciones_mensuales:
                ejecucion.monto_presupuestado = (
                    ejecucion.monto_presupuestado * ratio
                ).quantize(Decimal("0.01"))
                ejecucion.monto_disponible = (
                    ejecucion.monto_presupuestado - ejecucion.monto_ejecutado
                ).quantize(Decimal("0.01"))
                if ejecucion.monto_presupuestado > 0:
                    ejecucion.porcentaje_ejecucion = (
                        ejecucion.monto_ejecutado / ejecucion.monto_presupuestado * Decimal("100")
                    ).quantize(Decimal("0.01"))

    if "cuenta_nombre" in update_data and update_data["cuenta_nombre"] is not None:
        cuenta.cuenta_nombre = update_data["cuenta_nombre"]

    await _recalculate_cuenta(cuenta)
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="UPDATE",
        entity_type="presupuesto_cuenta",
        entity_id=cuenta.id,
        description=f"Cuenta presupuestaria actualizada: {cuenta.cuenta_codigo}",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoCuentaResponse.model_validate(cuenta)


@router.delete("/cuentas/{cuenta_id}")
async def delete_cuenta(
    cuenta_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Eliminar una cuenta presupuestaria (solo en borrador)"""
    result = await db.execute(
        select(PresupuestoCuenta).where(PresupuestoCuenta.id == cuenta_id)
    )
    cuenta = result.scalars().first()
    if not cuenta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cuenta presupuestaria no encontrada.",
        )

    presup_result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == cuenta.presupuesto_id)
    )
    presupuesto = presup_result.scalars().first()
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.BORRADOR.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden eliminar cuentas de presupuestos en borrador.",
        )

    await db.delete(cuenta)
    await db.flush()

    # Recalculate presupuesto totals
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="DELETE",
        entity_type="presupuesto_cuenta",
        entity_id=cuenta_id,
        description=f"Cuenta presupuestaria eliminada: {cuenta.cuenta_codigo}",
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Cuenta presupuestaria eliminada exitosamente."}


# ==========================================
# 11-13. Ejecución Mensual
# ==========================================

@router.post("/cuentas/{cuenta_id}/ejecucion", response_model=PresupuestoEjecucionMensualResponse, status_code=status.HTTP_201_CREATED)
async def register_ejecucion_mensual(
    cuenta_id: str,
    data: EjecucionMensualCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Registrar ejecución mensual para una cuenta presupuestaria"""
    result = await db.execute(
        select(PresupuestoCuenta).where(PresupuestoCuenta.id == cuenta_id)
    )
    cuenta = result.scalars().first()
    if not cuenta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cuenta presupuestaria no encontrada.",
        )

    # Get presupuesto to verify ownership
    presup_result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == cuenta.presupuesto_id)
    )
    presupuesto = presup_result.scalars().first()
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.APROBADO.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se puede registrar ejecución en presupuestos aprobados.",
        )

    # Determine current month (1-12)
    current_month = datetime.now(timezone.utc).month

    # Find ejecucion for current month
    ejec_result = await db.execute(
        select(PresupuestoEjecucionMensual).where(
            PresupuestoEjecucionMensual.presupuesto_cuenta_id == cuenta_id,
            PresupuestoEjecucionMensual.mes == current_month,
        )
    )
    ejecucion = ejec_result.scalars().first()

    if ejecucion:
        # Update existing ejecucion
        ejecucion.monto_ejecutado = data.monto_ejecutado
        ejecucion.monto_disponible = ejecucion.monto_presupuestado - data.monto_ejecutado
        if ejecucion.monto_presupuestado > 0:
            ejecucion.porcentaje_ejecucion = (
                data.monto_ejecutado / ejecucion.monto_presupuestado * Decimal("100")
            ).quantize(Decimal("0.01"))
        else:
            ejecucion.porcentaje_ejecucion = Decimal("0")
        if data.observaciones:
            ejecucion.observaciones = data.observaciones
    else:
        # Create new ejecucion
        monto_presupuestado = (cuenta.monto_anual / Decimal("12")).quantize(Decimal("0.01"))
        monto_disponible = monto_presupuestado - data.monto_ejecutado
        pct = (
            data.monto_ejecutado / monto_presupuestado * Decimal("100")
        ).quantize(Decimal("0.01")) if monto_presupuestado > 0 else Decimal("0")

        ejecucion = PresupuestoEjecucionMensual(
            presupuesto_cuenta_id=cuenta_id,
            mes=current_month,
            monto_presupuestado=monto_presupuestado,
            monto_ejecutado=data.monto_ejecutado,
            monto_disponible=monto_disponible,
            porcentaje_ejecucion=pct,
            observaciones=data.observaciones,
        )
        db.add(ejecucion)

    await db.flush()

    # Recalculate cuenta totals
    await _recalculate_cuenta(cuenta)

    # Check alerts
    await _check_alerts(db, cuenta, presupuesto.company_id)

    # Recalculate presupuesto totals
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        entity_type="ejecucion_mensual",
        entity_id=ejecucion.id,
        description=f"Ejecución mensual registrada: mes {ejecucion.mes}, ${data.monto_ejecutado}",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoEjecucionMensualResponse.model_validate(ejecucion)


@router.get("/cuentas/{cuenta_id}/ejecucion", response_model=list[PresupuestoEjecucionMensualResponse])
async def get_ejecucion_for_cuenta(
    cuenta_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Obtener ejecución mensual de una cuenta presupuestaria"""
    result = await db.execute(
        select(PresupuestoCuenta).where(PresupuestoCuenta.id == cuenta_id)
    )
    cuenta = result.scalars().first()
    if not cuenta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cuenta presupuestaria no encontrada.",
        )

    presup_result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == cuenta.presupuesto_id)
    )
    presupuesto = presup_result.scalars().first()
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    ejec_result = await db.execute(
        select(PresupuestoEjecucionMensual)
        .where(PresupuestoEjecucionMensual.presupuesto_cuenta_id == cuenta_id)
        .order_by(PresupuestoEjecucionMensual.mes)
    )
    ejecuciones = ejec_result.scalars().all()
    return [PresupuestoEjecucionMensualResponse.model_validate(e) for e in ejecuciones]


@router.put("/ejecucion/{ejecucion_id}", response_model=PresupuestoEjecucionMensualResponse)
async def update_ejecucion(
    ejecucion_id: str,
    data: EjecucionMensualUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Actualizar monto de una ejecución mensual"""
    result = await db.execute(
        select(PresupuestoEjecucionMensual).where(PresupuestoEjecucionMensual.id == ejecucion_id)
    )
    ejecucion = result.scalars().first()
    if not ejecucion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ejecución mensual no encontrada.",
        )

    # Get cuenta and presupuesto to verify ownership
    cuenta_result = await db.execute(
        select(PresupuestoCuenta).where(PresupuestoCuenta.id == ejecucion.presupuesto_cuenta_id)
    )
    cuenta = cuenta_result.scalars().first()

    presup_result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == cuenta.presupuesto_id)
    )
    presupuesto = presup_result.scalars().first()
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    if presupuesto.estado != PresupuestoEstado.APROBADO.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se puede modificar ejecución en presupuestos aprobados.",
        )

    # Update ejecucion
    ejecucion.monto_ejecutado = data.monto_ejecutado
    ejecucion.monto_disponible = ejecucion.monto_presupuestado - data.monto_ejecutado
    if ejecucion.monto_presupuestado > 0:
        ejecucion.porcentaje_ejecucion = (
            data.monto_ejecutado / ejecucion.monto_presupuestado * Decimal("100")
        ).quantize(Decimal("0.01"))
    else:
        ejecucion.porcentaje_ejecucion = Decimal("0")

    if data.observaciones is not None:
        ejecucion.observaciones = data.observaciones

    await db.flush()

    # Recalculate cuenta totals
    await _recalculate_cuenta(cuenta)

    # Check alerts
    await _check_alerts(db, cuenta, presupuesto.company_id)

    # Recalculate presupuesto totals
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="UPDATE",
        entity_type="ejecucion_mensual",
        entity_id=ejecucion.id,
        description=f"Ejecución mensual actualizada: mes {ejecucion.mes}, ${data.monto_ejecutado}",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoEjecucionMensualResponse.model_validate(ejecucion)


# ==========================================
# 14-17. Alertas (read/resolve)
# ==========================================

@router.put("/alertas/{alerta_id}/read", response_model=PresupuestoAlertaResponse)
async def mark_alerta_read(
    alerta_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Marcar una alerta como leída"""
    result = await db.execute(
        select(PresupuestoAlerta).where(PresupuestoAlerta.id == alerta_id)
    )
    alerta = result.scalars().first()
    if not alerta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alerta no encontrada.",
        )
    await _get_company_for_user(db, alerta.company_id, current_user.id)

    alerta.is_leida = True
    await db.flush()

    return PresupuestoAlertaResponse.model_validate(alerta)


@router.put("/alertas/{alerta_id}/resolve", response_model=PresupuestoAlertaResponse)
async def mark_alerta_resolved(
    alerta_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Marcar una alerta como resuelta"""
    result = await db.execute(
        select(PresupuestoAlerta).where(PresupuestoAlerta.id == alerta_id)
    )
    alerta = result.scalars().first()
    if not alerta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alerta no encontrada.",
        )
    await _get_company_for_user(db, alerta.company_id, current_user.id)

    alerta.is_leida = True
    alerta.is_resuelta = True
    await db.flush()

    return PresupuestoAlertaResponse.model_validate(alerta)


# ==========================================
# 21. Recalcular presupuesto
# ==========================================

@router.post("/{presupuesto_id}/recalcular", response_model=PresupuestoAnualResponse)
async def recalcular_presupuesto(
    presupuesto_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Recalcular ejecutado/disponible para todas las cuentas de un presupuesto"""
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    # Recalculate each cuenta
    for cuenta in presupuesto.cuentas:
        if not cuenta.is_active:
            continue
        await _recalculate_cuenta(cuenta)

        # Recalculate each ejecucion mensual
        for ejecucion in cuenta.ejecuciones_mensuales:
            ejecucion.monto_disponible = ejecucion.monto_presupuestado - ejecucion.monto_ejecutado
            if ejecucion.monto_presupuestado > 0:
                ejecucion.porcentaje_ejecucion = (
                    ejecucion.monto_ejecutado / ejecucion.monto_presupuestado * Decimal("100")
                ).quantize(Decimal("0.01"))
            else:
                ejecucion.porcentaje_ejecucion = Decimal("0")

        # Check alerts
        await _check_alerts(db, cuenta, presupuesto.company_id)

    # Recalculate presupuesto totals
    await _recalculate_presupuesto(presupuesto)
    await db.flush()

    await log_action(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="RECALCULATE",
        entity_type="presupuesto_anual",
        entity_id=presupuesto.id,
        description=f"Presupuesto recalculado: {presupuesto.nombre}",
        ip_address=request.client.host if request.client else None,
    )

    return PresupuestoAnualResponse.model_validate(presupuesto)


# ==========================================
# Feature 1: Budget Alerts (Fase 12)
# ==========================================

MESES_NOMBRES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


async def _get_ejecutado_from_asientos(
    db: AsyncSession,
    company_id: str,
    cuenta_codigo: str,
    anio: int,
    mes: int | None = None,
) -> Decimal:
    """
    Query actual spending from accounting entries (AsientoContable) for a budget account.
    Sums debit amounts for expense/cost accounts and credit amounts for revenue accounts.
    """
    from app.models.accounting import AsientoContable, AsientoDetalle, AsientoEstado, CuentaContable

    query = (
        select(
            func.coalesce(func.sum(AsientoDetalle.debito), Decimal("0"))
        )
        .join(AsientoContable, AsientoDetalle.asiento_id == AsientoContable.id)
        .join(CuentaContable, AsientoDetalle.cuenta_contable_id == CuentaContable.id)
        .where(
            AsientoContable.company_id == company_id,
            AsientoContable.estado == AsientoEstado.APROBADO.value,
            CuentaContable.codigo == cuenta_codigo,
            func.strftime("%Y", AsientoContable.fecha) == str(anio),
        )
    )

    if mes is not None:
        query = query.where(func.strftime("%m", AsientoContable.fecha) == f"{mes:02d}")

    result = await db.execute(query)
    return result.scalar() or Decimal("0")


@router.get("/alertas/realtime", response_model=BudgetAlertsResponse)
async def get_budget_alertas_realtime(
    company_id: str = Query(..., description="ID de la empresa"),
    anio: int | None = Query(None, description="Año fiscal"),
    nivel: str | None = Query(None, description="Filtrar por nivel: WARNING, CRITICAL, OVER"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get real-time budget alerts by comparing actual spending from accounting entries
    vs budgeted amounts. Alert levels:
    - WARNING: ejecutado >= 80% of presupuesto
    - CRITICAL: ejecutado >= 95% of presupuesto
    - OVER: ejecutado > 100% of presupuesto (sobregiro)
    """
    await _get_company_for_user(db, company_id, current_user.id)

    presupuesto_query = select(PresupuestoAnual).where(
        PresupuestoAnual.company_id == company_id,
        PresupuestoAnual.estado == PresupuestoEstado.APROBADO.value,
        PresupuestoAnual.is_active == True,
    )
    if anio is not None:
        presupuesto_query = presupuesto_query.where(PresupuestoAnual.anio == anio)

    result = await db.execute(presupuesto_query)
    presupuestos = result.scalars().all()

    alertas: list[BudgetAlertItem] = []

    for presupuesto in presupuestos:
        for cuenta in presupuesto.cuentas:
            if not cuenta.is_active or cuenta.cuenta_tipo != TipoCuenta.EGRESO.value:
                continue
            if cuenta.monto_anual <= 0:
                continue

            ejecutado = await _get_ejecutado_from_asientos(
                db, company_id, cuenta.cuenta_codigo, presupuesto.anio
            )

            porcentaje = (ejecutado / cuenta.monto_anual * Decimal("100")).quantize(Decimal("0.01"))
            diferencia = cuenta.monto_anual - ejecutado

            nivel_alerta = None
            if ejecutado > cuenta.monto_anual:
                nivel_alerta = "OVER"
            elif porcentaje >= Decimal("95"):
                nivel_alerta = "CRITICAL"
            elif porcentaje >= Decimal("80"):
                nivel_alerta = "WARNING"

            if nivel_alerta is None:
                continue

            if nivel is not None and nivel_alerta != nivel:
                continue

            alertas.append(BudgetAlertItem(
                cuenta_codigo=cuenta.cuenta_codigo,
                cuenta_nombre=cuenta.cuenta_nombre,
                cuenta_tipo=cuenta.cuenta_tipo,
                presupuesto=cuenta.monto_anual,
                ejecutado=ejecutado,
                porcentaje=porcentaje,
                nivel_alerta=nivel_alerta,
                diferencia=diferencia,
            ))

    severity_order = {"OVER": 0, "CRITICAL": 1, "WARNING": 2}
    alertas.sort(key=lambda a: (severity_order.get(a.nivel_alerta, 3), -a.porcentaje))

    return BudgetAlertsResponse(
        company_id=company_id,
        anio=anio,
        total_alertas=len(alertas),
        alertas=alertas,
    )


@router.get("/ejecucion/{presupuesto_id}", response_model=BudgetExecutionDetailResponse)
async def get_budget_ejecucion_detail(
    presupuesto_id: str,
    cuenta_codigo: str = Query(..., description="Código de cuenta contable"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Detailed budget execution report with month-by-month breakdown.
    Compares presupuesto_mes vs ejecutado_mes from actual accounting entries.
    Includes accumulated totals.
    """
    result = await db.execute(
        select(PresupuestoAnual).where(PresupuestoAnual.id == presupuesto_id)
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Presupuesto no encontrado.",
        )
    await _get_company_for_user(db, presupuesto.company_id, current_user.id)

    cuenta = None
    for c in presupuesto.cuentas:
        if c.cuenta_codigo == cuenta_codigo and c.is_active:
            cuenta = c
            break

    if not cuenta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Cuenta {cuenta_codigo} no encontrada en el presupuesto.",
        )

    detalle_mensual: list[BudgetExecutionMonthDetail] = []
    acumulado_presupuesto = Decimal("0")
    acumulado_ejecutado = Decimal("0")

    for mes_num in range(1, 13):
        presupuesto_mes = Decimal("0")
        for ejecucion in cuenta.ejecuciones_mensuales:
            if ejecucion.mes == mes_num:
                presupuesto_mes = ejecucion.monto_presupuestado
                break

        ejecutado_mes = await _get_ejecutado_from_asientos(
            db, presupuesto.company_id, cuenta_codigo, presupuesto.anio, mes_num
        )

        acumulado_presupuesto += presupuesto_mes
        acumulado_ejecutado += ejecutado_mes

        porcentaje_acumulado = (
            (acumulado_ejecutado / acumulado_presupuesto * Decimal("100")).quantize(Decimal("0.01"))
            if acumulado_presupuesto > 0
            else Decimal("0")
        )

        detalle_mensual.append(BudgetExecutionMonthDetail(
            mes=mes_num,
            mes_nombre=MESES_NOMBRES[mes_num - 1],
            presupuesto_mes=presupuesto_mes,
            ejecutado_mes=ejecutado_mes,
            diferencia_mes=presupuesto_mes - ejecutado_mes,
            acumulado_presupuesto=acumulado_presupuesto,
            acumulado_ejecutado=acumulado_ejecutado,
            porcentaje_acumulado=porcentaje_acumulado,
        ))

    total_ejecutado = sum(d.ejecutado_mes for d in detalle_mensual)
    porcentaje_ejecucion = (
        (total_ejecutado / cuenta.monto_anual * Decimal("100")).quantize(Decimal("0.01"))
        if cuenta.monto_anual > 0
        else Decimal("0")
    )

    return BudgetExecutionDetailResponse(
        presupuesto_id=presupuesto_id,
        presupuesto_nombre=presupuesto.nombre,
        anio=presupuesto.anio,
        cuenta_codigo=cuenta_codigo,
        cuenta_nombre=cuenta.cuenta_nombre,
        cuenta_tipo=cuenta.cuenta_tipo,
        total_presupuesto=cuenta.monto_anual,
        total_ejecutado=total_ejecutado,
        porcentaje_ejecucion=porcentaje_ejecucion,
        detalle_mensual=detalle_mensual,
    )


# ==========================================
# Budget Excel Export (Fase 12)
# ==========================================

try:
    import io as _io
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import (
        Font as _Font,
        Alignment as _Alignment,
        PatternFill as _PatternFill,
        Border as _Border,
        Side as _Side,
    )
    from fastapi.responses import StreamingResponse as _StreamingResponse
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@router.get("/export/excel")
async def export_budget_to_excel(
    company_id: str = Query(..., description="ID de la empresa"),
    anio: int = Query(..., description="Año fiscal"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Export budget execution to Excel file.
    Returns an Excel file with budget vs actual comparison for all accounts.
    """
    if not HAS_OPENPYXL:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="La exportacion a Excel requiere el paquete 'openpyxl'. Instalelo con: pip install openpyxl",
        )

    await _get_company_for_user(db, company_id, current_user.id)

    result = await db.execute(
        select(PresupuestoAnual).where(
            PresupuestoAnual.company_id == company_id,
            PresupuestoAnual.anio == anio,
            PresupuestoAnual.is_active == True,
        )
    )
    presupuesto = result.scalars().first()
    if not presupuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontro presupuesto activo para el anio {anio}.",
        )

    wb = _Workbook()
    ws = wb.active
    ws.title = f"Presupuesto {anio}"

    header_font = _Font(bold=True, color="FFFFFF", size=11)
    header_fill = _PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = _Font(bold=True, size=14, color="2F5496")
    warning_fill = _PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    critical_fill = _PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    over_fill = _PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = _Border(
        left=_Side(style="thin"),
        right=_Side(style="thin"),
        top=_Side(style="thin"),
        bottom=_Side(style="thin"),
    )
    number_format = "#,##0.00"
    percent_format = "0.00%"

    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Presupuesto vs Ejecucion {anio} - {presupuesto.nombre}"
    title_cell.font = title_font
    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = _Font(italic=True, size=10)

    headers = [
        "Codigo", "Cuenta", "Tipo", "Presupuesto", "Ejecutado",
        "Diferencia", "% Ejecucion", "Nivel Alerta", "Presup. Mensual", "Ejecutado Mensual",
    ]
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = _Alignment(horizontal="center")
        cell.border = thin_border

    row_num = header_row + 1
    for cuenta in presupuesto.cuentas:
        if not cuenta.is_active:
            continue

        ejecutado = await _get_ejecutado_from_asientos(
            db, company_id, cuenta.cuenta_codigo, anio
        )
        diferencia = cuenta.monto_anual - ejecutado
        porcentaje = (
            ejecutado / cuenta.monto_anual if cuenta.monto_anual > 0 else Decimal("0")
        )

        nivel_alerta = ""
        row_fill = None
        if ejecutado > cuenta.monto_anual:
            nivel_alerta = "OVER"
            row_fill = over_fill
        elif porcentaje >= Decimal("0.95"):
            nivel_alerta = "CRITICAL"
            row_fill = critical_fill
        elif porcentaje >= Decimal("0.80"):
            nivel_alerta = "WARNING"
            row_fill = warning_fill

        presupuesto_mensual = ""
        for ejecucion in cuenta.ejecuciones_mensuales:
            presupuesto_mensual += f"M{ejecucion.mes}: ${ejecucion.monto_presupuestado:,.2f}\n"

        row_data = [
            cuenta.cuenta_codigo,
            cuenta.cuenta_nombre,
            cuenta.cuenta_tipo,
            float(cuenta.monto_anual),
            float(ejecutado),
            float(diferencia),
            float(porcentaje),
            nivel_alerta,
            presupuesto_mensual.strip(),
            "",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border

            if col_num in (4, 5, 6):
                cell.number_format = number_format
            elif col_num == 7:
                cell.number_format = percent_format

            if row_fill and col_num <= 8:
                cell.fill = row_fill

        row_num += 1

    row_num += 2
    ws.merge_cells(f"A{row_num}:C{row_num}")
    ws.cell(row=row_num, column=1).value = "RESUMEN"
    ws.cell(row=row_num, column=1).font = _Font(bold=True, size=12)

    row_num += 1
    summary_items = [
        ("Total Ingresos Presupuestado", float(presupuesto.total_ingresos_presupuestado)),
        ("Total Ingresos Ejecutado", float(presupuesto.total_ingresos_ejecutado)),
        ("Total Egresos Presupuestado", float(presupuesto.total_egresos_presupuestado)),
        ("Total Egresos Ejecutado", float(presupuesto.total_egresos_ejecutado)),
        ("Resultado Presupuestario", float(presupuesto.total_ingresos_presupuestado - presupuesto.total_egresos_presupuestado)),
        ("Resultado Real", float(presupuesto.total_ingresos_ejecutado - presupuesto.total_egresos_ejecutado)),
    ]

    for label, value in summary_items:
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=1).font = _Font(bold=True)
        ws.cell(row=row_num, column=4).value = value
        ws.cell(row=row_num, column=4).number_format = number_format
        row_num += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    for col_letter in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 20

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"presupuesto_{anio}_{company_id[:8]}.xlsx"

    return _StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
